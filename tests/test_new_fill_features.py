import openpyxl
import pytest
import pandas as pd
from datetime import datetime, time, timedelta
from src.cleaning import _parse_shift_movements_table, CraneRow
from src.merge import CraneSession, MergedVesselDataset, BreakBulkEntry
from src.report_generator import (
    TPFREPWriter,
    fill_crane_delays_section,
    fill_restow_sections,
    fill_hatch_cover_section,
    fill_break_bulk_section
)

def test_parse_shift_movements_table():
    rows = [
        [None, None, None],
        [None, None, None],
        [None, None, None],
        [None, None, None],
        [None, None, None],
        [None, None, None],
        [None, None, None],
        [None, None, None],
        [None, None, None],
        [None, None, None],
        [None, None, None],
        [None, None, None],
        # Row 13 (0-indexed 12)
        ["Portiques", None, "P2", "P3", "P4"],
        ["Import", None, 10, 20, 30],
        ["Export", None, 0, 0, 0],
        ["Shifting déchargé", None, 1, 0, 2],
        ["Shifting chargé", None, 0, 2, 0],
        ["Ouverture P/C", None, 3, 1, 0],
        ["Fermeture P/C", None, 2, 0, 4]
    ]
    
    import config
    orig_cranes = config.CRANE_IDS
    try:
        config.CRANE_IDS = ["P1", "P2", "P3", "P4"]
        res = _parse_shift_movements_table(rows)
        
        assert res["P2"]["shifting_discharged"] == 1
        assert res["P3"]["shifting_loaded"] == 2
        assert res["P2"]["hatch_open"] == 3
        assert res["P4"]["hatch_close"] == 4
    finally:
        config.CRANE_IDS = orig_cranes

def test_fill_crane_delays_section():
    wb = openpyxl.Workbook()
    ws = wb.active
    
    blue_fill = openpyxl.styles.PatternFill(
        fill_type="solid", fgColor=openpyxl.styles.colors.Color(indexed=44)
    )
    
    # row 8: crane timesheet anchor
    ws.cell(row=8, column=3, value="crane timesheet")
    # row 10: crane IDs
    ws.cell(row=10, column=5, value="P2").fill = blue_fill
    ws.cell(row=10, column=6, value="P3").fill = blue_fill
    # row 15: section anchor for the port/terminal delays section
    ws.cell(row=15, column=3, value="delays caused by port or terminal")
    # rows 18 and 19: consecutive blue cells for two delay slots (P2 col=5, P3 col=6)
    ws.cell(row=18, column=5).fill = blue_fill
    ws.cell(row=18, column=6).fill = blue_fill
    ws.cell(row=19, column=5).fill = blue_fill
    ws.cell(row=19, column=6).fill = blue_fill
    
    writer = TPFREPWriter(ws)
    
    from src.cleaning import GeneralDelayEntry
    merged = MergedVesselDataset(
        vessel_name="BELITAKI",
        escale="0XP93W1MA",
        crane_sessions={
            "P2": [
                CraneSession(
                    shift_num=1,
                    commenced=datetime(2026, 6, 21, 8, 0),
                    completed=datetime(2026, 6, 21, 12, 0),
                    import_moves=10,
                    export_moves=5
                )
            ],
            "P3": [
                CraneSession(
                    shift_num=1,
                    commenced=datetime(2026, 6, 21, 8, 0),
                    completed=datetime(2026, 6, 21, 12, 0),
                    import_moves=10,
                    export_moves=5
                )
            ]
        },
        general_delays=[
            GeneralDelayEntry(
                shift_num=1,
                label="P2",
                duration_minutes=30.0,
                reason_text="vent fort"
            ),
            GeneralDelayEntry(
                shift_num=1,
                label="P2",
                duration_minutes=45.0,
                reason_text="attente lashing navire"
            )
        ]
    )
    
    import config
    orig_cranes = config.CRANE_IDS
    try:
        config.CRANE_IDS = ["P1", "P2", "P3", "P4"]
        fill_crane_delays_section(writer, merged)
        
        # First delay slot: row 18, col 5 (P2)
        assert ws.cell(row=18, column=5).value is not None, (
            f"Row 18 col 5 attendu non-None, obtenu {ws.cell(row=18, column=5).value}"
        )
        # Second delay slot: row 19, col 5 (P2)
        assert ws.cell(row=19, column=5).value is not None, (
            f"Row 19 col 5 attendu non-None, obtenu {ws.cell(row=19, column=5).value}"
        )
    finally:
        config.CRANE_IDS = orig_cranes

def test_fill_restow_sections():
    wb = openpyxl.Workbook()
    ws = wb.active
    
    blue_fill = openpyxl.styles.PatternFill(
        fill_type="solid", fgColor=openpyxl.styles.colors.Color(indexed=44)
    )
    
    ws.cell(row=201, column=3, value="5.1 Discharge+Reload (two moves)")
    ws.cell(row=202, column=4, value="Full")
    ws.cell(row=202, column=7, value="Empty")
    ws.cell(row=203, column=4, value="20'")
    ws.cell(row=203, column=5, value="40'  +  45'")
    ws.cell(row=203, column=7, value="20'")
    ws.cell(row=203, column=8, value="40'  +  45'")
    
    ws.cell(row=204, column=4).fill = blue_fill
    ws.cell(row=204, column=5).fill = blue_fill
    ws.cell(row=204, column=7).fill = blue_fill
    ws.cell(row=204, column=8).fill = blue_fill
    
    writer = TPFREPWriter(ws)
    
    merged = MergedVesselDataset(
        vessel_name="BELITAKI",
        escale="0XP93W1MA",
        containers_import=pd.DataFrame({"operator": ["CMA"]}),
        containers_export=pd.DataFrame({"operator": []}),
        crane_sessions={
            "P2": [
                CraneSession(
                    shift_num=1,
                    commenced=datetime(2026, 6, 21, 8, 0),
                    completed=datetime(2026, 6, 21, 12, 0),
                    import_moves=10,
                    export_moves=5,
                    restow_discharged=5,
                    restow_loaded=0
                )
            ]
        }
    )
    
    fill_restow_sections(writer, merged)
    assert ws.cell(row=204, column=5).value == 5

def test_fill_hatch_cover_section():
    wb = openpyxl.Workbook()
    ws = wb.active
    
    blue_fill = openpyxl.styles.PatternFill(
        fill_type="solid", fgColor=openpyxl.styles.colors.Color(indexed=44)
    )
    
    ws.cell(row=231, column=3, value="6 Hatch Cover Moves")
    ws.cell(row=233, column=3, value="Total")
    ws.cell(row=233, column=4).fill = blue_fill
    
    writer = TPFREPWriter(ws)
    merged = MergedVesselDataset(
        vessel_name="BELITAKI",
        escale="0XP93W1MA",
        crane_sessions={
            "P2": [
                CraneSession(
                    shift_num=1,
                    commenced=datetime(2026, 6, 21, 8, 0),
                    completed=datetime(2026, 6, 21, 12, 0),
                    import_moves=10,
                    export_moves=5,
                    hatch_cover_open=2,
                    hatch_cover_close=4
                )
            ]
        }
    )
    
    fill_hatch_cover_section(writer, merged)
    assert ws.cell(row=233, column=4).value == 6

def test_fill_break_bulk_section():
    wb = openpyxl.Workbook()
    ws = wb.active
    
    blue_fill = openpyxl.styles.PatternFill(
        fill_type="solid", fgColor=openpyxl.styles.colors.Color(indexed=44)
    )
    
    ws.cell(row=236, column=3, value="7 Break Bulk Moves")
    ws.cell(row=238, column=4, value="Load BB")
    ws.cell(row=238, column=5, value="Discharge BB")
    ws.cell(row=239, column=3, value="CMA")
    ws.cell(row=239, column=4).fill = blue_fill
    ws.cell(row=239, column=5).fill = blue_fill
    
    writer = TPFREPWriter(ws)
    merged = MergedVesselDataset(
        vessel_name="BELITAKI",
        escale="0XP93W1MA",
        containers_import=pd.DataFrame({"operator": ["CMA"]}),
        containers_export=pd.DataFrame({"operator": []}),
        break_bulk_moves=[
            BreakBulkEntry(
                operator="CMA",
                is_load=False,
                is_discharge=True,
                count=1
            )
        ]
    )
    
    fill_break_bulk_section(writer, merged)
    assert ws.cell(row=239, column=5).value == 1
    assert ws.cell(row=239, column=4).value == 0
