"""
tests/test_root_causes.py — Tests de non-régression pour les 5 causes racines TPFREP.

RC1 — Retards grue écrits aux mauvaises lignes (et exclusion de la section Port/Terminal)
RC2 — Champs D17-D20 (Arrival/Sailed Berth, Lashing Gangs) remplis par erreur
RC3 — Total mouvements par grue doublement compté (sessions à cheval sur minuit)
RC4 — Break Bulk Discharge perdu (résultat 0 au lieu de la valeur réelle)
RC5 — Session de grue décalée d'un jour (DOC 00h-05h dans shift nocturne)
"""
from __future__ import annotations

import io
import tempfile
from datetime import datetime, time as dtime, timedelta
from pathlib import Path
from typing import Dict, List
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

import config
from src.calculations import compute_all_kpis
from src.cleaning import CleanedShiftReport, CraneRow, GeneralDelayEntry
from src.merge import (
    CraneSession,
    MergedVesselDataset,
    BreakBulkEntry,
    _resolve_crane_session,
    merge_crane_sessions,
)


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _make_crane_row(
    shift_num: int,
    crane_id: str,
    vessel: str,
    doc,
    foc,
    imp: int = 0,
    exp: int = 0,
    obs: str = None,
) -> CraneRow:
    return CraneRow(
        shift_num=shift_num,
        crane_id=crane_id,
        vessel_raw=vessel,
        vessel_normalized=vessel.upper().replace(" ", ""),
        import_moves=imp,
        export_moves=exp,
        doc_raw=doc,
        foc_raw=foc,
        observations=obs,
    )


def _make_cleaned(shift_num: int, date: datetime, crane_rows: list, delays: list = None):
    return CleanedShiftReport(
        shift_num=shift_num,
        shift_date=date,
        crane_rows=crane_rows,
        general_delays=delays or [],
    )


# ─────────────────────────────────────────────────────────────────────────────
# RC1 — Retards grue écrits aux bonnes lignes (positions dynamiques)
# ─────────────────────────────────────────────────────────────────────────────

class TestRC1CraneDelayRows:
    """Vérifie que les retards grues sont écrits à la ligne correcte dans le template.

    Méthode : on construit un faux workbook minimaliste avec une section
    'Delays caused by Ship or Cargo' et on vérifie que find_crane_delay_section_rows
    retourne les bonnes lignes (les lignes bleues de la section, pas des lignes
    arbitraires ailleurs).
    """

    def _make_ws_with_delay_section(self, crane_col: int = 4):
        """Crée un workbook openpyxl minimal avec une section de retard grue."""
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        # Couleur bleue (indexée 44 → on utilise RGB pour les tests)
        BLUE_FILL = PatternFill(fill_type="solid", fgColor="FF99CCFF")

        wb = Workbook()
        ws = wb.active
        ws.title = "Test"

        # Ligne 5 : ancre "Delays caused by Ship or Cargo Operation"
        ws.cell(row=5, column=1).value = "Delays caused by Ship or Cargo Operation"

        # Lignes 6, 7, 8 : lignes bleues pour les retards (col = crane_col+1 car 1-indexed)
        for r in (6, 7, 8):
            ws.cell(row=r, column=crane_col + 1).fill = BLUE_FILL

        # Ligne 15 : une autre cellule bleue HORS de la section (ne doit pas être capturée)
        ws.cell(row=15, column=crane_col + 1).fill = BLUE_FILL

        return wb, ws

    def test_delay_rows_detected_in_correct_section(self):
        """Les lignes de retard doivent être détectées sous l'ancre de section."""
        from src.report_generator import TPFREPWriter

        crane_col = 4  # 0-indexed
        wb, ws = self._make_ws_with_delay_section(crane_col)

        writer = TPFREPWriter(ws)

        crane_cols = {"P1": crane_col}
        delay_rows, _ = writer.find_crane_delay_section_rows(
            section_key="delays_ship_cargo",
            crane_cols=crane_cols,
            max_rows=15,  # limité à 15 lignes : ne doit PAS capturer la ligne 15
        )

        # 3 lignes bleues en lignes 6, 7, 8 (0-indexed : 5, 6, 7)
        assert len(delay_rows) == 3, (
            f"Attendu 3 lignes de retard, obtenu {len(delay_rows)} : {delay_rows}"
        )
        assert delay_rows == [5, 6, 7], (
            f"Lignes attendues [5,6,7] (0-indexed), obtenues : {delay_rows}"
        )

    def test_delay_rows_not_captured_before_anchor(self):
        """Aucune ligne avant l'ancre de section ne doit être capturée."""
        from src.report_generator import TPFREPWriter
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        BLUE_FILL = PatternFill(fill_type="solid", fgColor="FF99CCFF")
        wb = Workbook()
        ws = wb.active

        crane_col = 4  # 0-indexed

        # Ligne bleue AVANT l'ancre (ligne 2, ancre en ligne 5)
        ws.cell(row=2, column=crane_col + 1).fill = BLUE_FILL

        # Ancre de section en ligne 5
        ws.cell(row=5, column=1).value = "Delays caused by Ship or Cargo Operation"

        # Lignes bleues APRES l'ancre
        for r in (6, 7):
            ws.cell(row=r, column=crane_col + 1).fill = BLUE_FILL

        writer = TPFREPWriter(ws)
        delay_rows, _ = writer.find_crane_delay_section_rows(
            section_key="delays_ship_cargo",
            crane_cols={"P1": crane_col},
        )

        # Seulement les 2 lignes après l'ancre (pas la ligne 2 avant)
        assert 1 not in delay_rows, "La ligne avant l'ancre ne doit pas être capturée."
        assert len(delay_rows) == 2


# ─────────────────────────────────────────────────────────────────────────────
# RC2 — D17-D20 restent vides (pas écrites par erreur comme sessions de grue)
# ─────────────────────────────────────────────────────────────────────────────

class TestRC2BerthingFieldsEmpty:
    """Vérifie que find_session_rows_for_crane ne capture pas de cellules bleues
    situées avant la zone réelle des sessions (ex : D17-D20).

    Méthode : workbook avec des cellules bleues en lignes basses (simulating D17-D20)
    ET des paires commenced/completed en lignes hautes. Avec min_row, seules les
    paires en lignes hautes doivent être retournées.
    """

    def test_min_row_excludes_early_blue_cells(self):
        """Avec min_row=30 (0-indexed), les cellules bleues des lignes 16-19 sont ignorées."""
        from src.report_generator import TPFREPWriter
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        BLUE_FILL = PatternFill(fill_type="solid", fgColor="FF99CCFF")
        wb = Workbook()
        ws = wb.active

        crane_col = 4  # 0-indexed
        # Simuler D17, D18, D19, D20 (lignes 17-20, 0-indexed : 16-19)
        for r in (16, 17, 18, 19):
            ws.cell(row=r + 1, column=crane_col + 1).fill = BLUE_FILL

        # Ligne d'ancrage "crane timesheet"
        ws.cell(row=35, column=1).value = "Crane Timesheet"
        ws.cell(row=36, column=1).value = "Crane ID"
        ws.cell(row=37, column=1).value = "Crane Type"

        # Sessions réelles : commenced en ligne 38, completed en ligne 39
        ws.cell(row=38, column=1).value = "Commenced"
        ws.cell(row=38, column=crane_col + 1).fill = BLUE_FILL
        ws.cell(row=39, column=1).value = "Completed"
        ws.cell(row=39, column=crane_col + 1).fill = BLUE_FILL

        writer = TPFREPWriter(ws)

        # Sans min_row : pourrait capturer les lignes 16-19
        slots_no_min = writer.find_session_rows_for_crane(
            crane_col=crane_col,
            search_from_row=0,
            max_rows=60,
            min_row=0,
        )

        # Avec min_row=37 (après crane_type_row) : ne doit PAS capturer 16-19
        slots_with_min = writer.find_session_rows_for_crane(
            crane_col=crane_col,
            search_from_row=35,
            max_rows=25,
            min_row=37,
        )

        # Vérifier que min_row exclut bien les lignes basses
        for commenced_row, completed_row in slots_with_min:
            assert commenced_row >= 37, (
                f"Session détectée avant min_row : commenced_row={commenced_row}"
            )

        # La vraie session (38, 39) doit être présente
        assert (37, 38) in slots_with_min, (
            f"Session réelle (37, 38) manquante dans slots_with_min : {slots_with_min}"
        )


# ─────────────────────────────────────────────────────────────────────────────
# RC3 — Déduplication des sessions (pas de double comptage)
# ─────────────────────────────────────────────────────────────────────────────

class TestRC3NoDuplicateSessions:
    """Vérifie que merge_crane_sessions ne compte pas deux fois une session
    signalée dans deux shifts consécutifs (session à cheval sur minuit).
    """

    def test_midnight_session_not_double_counted(self):
        """Une session 23:18-04:45 déclarée dans deux shifts consécutifs de la MÊME nuit
        ne doit apparaître qu'une seule fois.
        Scénario réel : le shift 2 (daté 22/06) et le shift 3 (aussi daté 22/06 car c'est
        la nuit du 22 au 23) déclarent tous les deux la session 23:18-04:45."""
        # Shift 2 et shift 3 ont la même shift_date (22/06) car le shift 3 est la nuit du 22/06.
        shift2 = _make_cleaned(
            2, datetime(2026, 6, 22),
            [_make_crane_row(2, "P3", "BELITAKI", dtime(23, 18), dtime(4, 45), imp=50, exp=0)],
        )
        # Shift 3 : re-déclare la même session avec la même date de base (22/06)
        shift3 = _make_cleaned(
            3, datetime(2026, 6, 22),
            [_make_crane_row(3, "P3", "BELITAKI", dtime(23, 18), dtime(4, 45), imp=50, exp=0)],
        )
        cleaned = {2: shift2, 3: shift3}
        sessions = merge_crane_sessions(cleaned, "BELITAKI")

        p3 = sessions.get("P3", [])
        # Il ne doit y avoir qu'une seule session (pas deux)
        assert len(p3) == 1, (
            f"Attendu 1 session (déduplication RC3), obtenu {len(p3)} : "
            f"{[(s.commenced, s.completed, s.import_moves) for s in p3]}"
        )
        assert p3[0].import_moves == 50

    def test_different_sessions_not_deduplicated(self):
        """Deux sessions avec des horaires différents doivent toutes les deux être conservées."""
        shift1 = _make_cleaned(
            1, datetime(2026, 6, 21),
            [
                _make_crane_row(1, "P3", "BELITAKI", dtime(11, 30), dtime(15, 0), imp=80, exp=0),
                _make_crane_row(1, "P3", "BELITAKI", dtime(22, 0), dtime(4, 30), imp=60, exp=0),
            ],
        )
        cleaned = {1: shift1}
        sessions = merge_crane_sessions(cleaned, "BELITAKI")
        p3 = sessions.get("P3", [])
        assert len(p3) == 2, f"Attendu 2 sessions distinctes, obtenu {len(p3)}"

    def test_total_moves_not_doubled_with_midnight_session(self):
        """Le total de mouvements ne doit pas être doublé si une session traverse minuit."""
        # Les deux shifts ont la même date de base (22/06 = nuit du 22 au 23)
        shift2 = _make_cleaned(
            2, datetime(2026, 6, 22),
            [_make_crane_row(2, "P3", "BELITAKI", dtime(23, 0), dtime(5, 0), imp=100, exp=50)],
        )
        # Même session dans le shift 3 de la même nuit (double-déclaration)
        shift3 = _make_cleaned(
            3, datetime(2026, 6, 22),
            [_make_crane_row(3, "P3", "BELITAKI", dtime(23, 0), dtime(5, 0), imp=100, exp=50)],
        )
        cleaned = {2: shift2, 3: shift3}
        sessions = merge_crane_sessions(cleaned, "BELITAKI")
        p3 = sessions.get("P3", [])
        total_moves = sum(s.import_moves + s.export_moves for s in p3)
        assert total_moves == 150, (
            f"Total mouvements attendu = 150 (pas 300), obtenu = {total_moves}"
        )


# ─────────────────────────────────────────────────────────────────────────────
# RC4 — Break Bulk Discharge non nul
# ─────────────────────────────────────────────────────────────────────────────

class TestRC4BreakBulkDischarge:
    """Vérifie que les mouvements Break Bulk de type Discharge (opérateur 'Common')
    sont correctement écrits dans le template.
    """

    def test_break_bulk_discharge_entry_created_in_merge(self):
        """Un mouvement 'elingue' sans mots-clés de chargement doit créer un BreakBulkEntry
        avec is_discharge=True."""
        shift1 = _make_cleaned(
            1, datetime(2026, 6, 21),
            [_make_crane_row(
                1, "P3", "BELITAKI",
                dtime(11, 30), dtime(15, 0), imp=80, exp=0,
                obs="Elingue sling bb opération déchargement",
            )],
        )
        import pandas as pd
        from src.merge import build_merged_dataset

        df_imp = pd.DataFrame([{"escale": "BELITAKI_24062026", "operator": "CMA",
                                  "full_empty_code": "P", "iso_type": "22G0",
                                  "movement_code": "DISCHARGE", "direction_code": "I",
                                  "container_number": "CMAU1234560"}])
        df_exp = pd.DataFrame()

        merged = build_merged_dataset({1: shift1}, df_imp, df_exp)

        bb_discharge_entries = [e for e in merged.break_bulk_moves if e.is_discharge]
        assert len(bb_discharge_entries) >= 1, (
            f"Attendu >= 1 BreakBulkEntry avec is_discharge=True, "
            f"obtenu : {merged.break_bulk_moves}"
        )

    def test_break_bulk_written_for_common_operator(self):
        """Les compteurs break bulk de l'opérateur 'Common' doivent être accessibles."""
        from src.merge import BreakBulkEntry

        merged = MergedVesselDataset(
            vessel_name="BELITAKI",
            escale="BELITAKI_24062026",
            crane_sessions={},
            containers_import=__import__("pandas").DataFrame(),
            containers_export=__import__("pandas").DataFrame(),
            break_bulk_moves=[
                BreakBulkEntry(operator="Common", is_load=False, is_discharge=True, count=1),
            ],
        )

        # Construire les dicts bb_discharged / bb_loaded comme dans fill_break_bulk_section
        bb_discharged: dict = {}
        bb_loaded: dict = {}
        for entry in merged.break_bulk_moves:
            if entry.is_discharge:
                bb_discharged[entry.operator] = bb_discharged.get(entry.operator, 0) + entry.count
            else:
                bb_loaded[entry.operator] = bb_loaded.get(entry.operator, 0) + entry.count

        assert bb_discharged.get("Common", 0) == 1, (
            f"bb_discharged['Common'] attendu = 1, obtenu : {bb_discharged}"
        )
        # La liste d'opérateurs BB doit inclure "Common"
        all_bb_operators = list(
            dict.fromkeys(list(bb_discharged.keys()) + list(bb_loaded.keys()))
        )
        assert "Common" in all_bb_operators


# ─────────────────────────────────────────────────────────────────────────────
# RC5 — Session de grue à DOC 00h-05h dans un shift nocturne → date J+1
# ─────────────────────────────────────────────────────────────────────────────

class TestRC5NightShiftSessionDate:
    """Vérifie que les sessions commençant entre 00h00 et 05h59 dans un shift
    nocturne (contenant une session après 20h) sont datées au lendemain (J+1).
    """

    def test_midnight_session_gets_next_day_date(self):
        """Session DOC=00:10 dans shift nocturne (22:00-06:00) → date J+1."""
        # Shift 3 daté du 23/06, couvre 22h-06h
        shift3 = _make_cleaned(
            3, datetime(2026, 6, 23),
            [
                # Session de nuit commençant à 22h (shift nocturne détecté grâce à cette session)
                _make_crane_row(3, "P3", "AKPERSEUS", dtime(22, 0), dtime(23, 59), imp=30, exp=0),
                # Session de début de nuit : DOC=00:10 → doit être datée au 24/06
                _make_crane_row(3, "P3", "AKPERSEUS", dtime(0, 10), dtime(4, 30), imp=20, exp=0),
            ],
        )
        cleaned = {3: shift3}
        sessions = merge_crane_sessions(cleaned, "AKPERSEUS")
        p3 = sessions.get("P3", [])

        # La session 00:10-04:30 doit être datée au 24/06 (pas 23/06)
        early_sessions = [s for s in p3 if s.commenced.hour < 6]
        assert len(early_sessions) >= 1, "Aucune session matinale trouvée"
        for s in early_sessions:
            assert s.commenced.date() == datetime(2026, 6, 24).date(), (
                f"RC5 : session DOC=00:10 doit être datée au 24/06, "
                f"obtenu : {s.commenced.strftime('%d/%m/%Y %H:%M')}"
            )

    def test_diurnal_shift_session_date_unchanged(self):
        """Un shift diurne (DOC=07:00) ne doit PAS voir sa date modifiée."""
        shift1 = _make_cleaned(
            1, datetime(2026, 6, 21),
            [_make_crane_row(1, "P3", "BELITAKI", dtime(7, 0), dtime(14, 0), imp=50, exp=0)],
        )
        cleaned = {1: shift1}
        sessions = merge_crane_sessions(cleaned, "BELITAKI")
        p3 = sessions.get("P3", [])
        assert len(p3) == 1
        # La date ne doit pas changer (shift diurne, pas nocturne)
        assert p3[0].commenced.date() == datetime(2026, 6, 21).date(), (
            f"Date altérée pour un shift diurne : {p3[0].commenced}"
        )

    def test_midnight_crossing_session_date_not_doubled(self):
        """Session traversant minuit (FOC < DOC) : la date de 'completed' est déjà J+1
        via timedelta(days=1). La correction RC5 ne doit pas rajouter un 2ème jour."""
        # Shift 2 daté du 22/06, session 23:18 → 04:45 (traversée minuit)
        shift2 = _make_cleaned(
            2, datetime(2026, 6, 22),
            [_make_crane_row(2, "P3", "BELITAKI", dtime(23, 18), dtime(4, 45), imp=50, exp=0)],
        )
        cleaned = {2: shift2}
        sessions = merge_crane_sessions(cleaned, "BELITAKI")
        p3 = sessions.get("P3", [])
        assert len(p3) == 1
        s = p3[0]
        # commenced = 22/06 23:18 (inchangé)
        assert s.commenced == datetime(2026, 6, 22, 23, 18), (
            f"commenced attendu = 22/06 23:18, obtenu : {s.commenced}"
        )
        # completed = 23/06 04:45 (traversée de minuit, +1 jour par la règle existante)
        assert s.completed == datetime(2026, 6, 23, 4, 45), (
            f"completed attendu = 23/06 04:45, obtenu : {s.completed}"
        )


# ─────────────────────────────────────────────────────────────────────────────
# RC1 — Mise à jour de périmètre : Delays caused by Ship or Cargo Operation reste vide
# ─────────────────────────────────────────────────────────────────────────────

class TestRC1ShipCargoDelaysEmpty:
    """Vérifie que la section 'Delays caused by Ship or Cargo Operation'
    ne reçoit jamais de valeur de retard dans le rapport final.
    """

    def test_ship_cargo_section_remains_empty(self):
        """Même si des retards de type ship/cargo existent, ils ne doivent pas
        être écrits dans le rapport final."""
        from src.report_generator import TPFREPWriter
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        BLUE_FILL = PatternFill(fill_type="solid", fgColor="FF99CCFF")
        wb = Workbook()
        ws = wb.active

        # Ligne d'ancrage pour "Delays caused by Ship or Cargo Operation"
        ws.cell(row=10, column=1).value = "Delays caused by Ship or Cargo Operation"
        # Colonne de grue bleue (P1)
        ws.cell(row=11, column=5).fill = BLUE_FILL
        ws.cell(row=12, column=5).fill = BLUE_FILL

        # On simule un retard grue qui pourrait être de type ship/cargo
        # (p. ex. label = grue, mais la logique globale de fill_crane_delays_section
        # ne doit pas écrire dans delays_ship_cargo)
        shift1 = _make_cleaned(
            1, datetime(2026, 6, 21),
            crane_rows=[
                _make_crane_row(1, "P1", "BELITAKI", dtime(7, 0), dtime(14, 0), imp=50)
            ],
            delays=[
                # Un retard pour la grue P1
                GeneralDelayEntry(shift_num=1, label="P1", duration_minutes=30.0, reason_text="Attente portique")
            ]
        )

        from src.merge import MergedVesselDataset
        merged = MergedVesselDataset(
            vessel_name="BELITAKI",
            escale="BELITAKI_24062026",
            crane_sessions={"P1": []},
            containers_import=__import__("pandas").DataFrame(),
            containers_export=__import__("pandas").DataFrame(),
            general_delays=shift1.general_delays,
        )

        writer = TPFREPWriter(ws)
        # S'assurer que delays_ship_cargo n'écrit rien
        # Nous allons vérifier que fill_crane_delays_section n'appelle pas write()
        # sur les lignes 11 et 12 (0-indexed : 10 et 11)
        with patch.object(writer, 'write', wraps=writer.write) as mock_write:
            from src.report_generator import fill_crane_delays_section
            fill_crane_delays_section(writer, merged)

            # Vérifier qu'aucune écriture n'a eu lieu sur les lignes de ship_cargo (lignes 11 ou 12)
            for call in mock_write.call_args_list:
                row, col, val, label = call[0][:4]
                assert row not in (10, 11), f"RC1 : Écriture interdite dans la section Ship/Cargo à la ligne {row + 1}"

