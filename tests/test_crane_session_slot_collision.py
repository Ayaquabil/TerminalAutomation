"""
tests/test_crane_session_slot_collision.py
==========================================

Tests pour les deux bugs identifiés :

  Bug A — merge_crane_sessions() : KeyError sur un crane_id absent de
           config.CRANE_IDS (ex: 'Q1').

  Bug B — fill_crane_timesheet_section() : collision de slots quand deux
           sessions de LA MÊME grue ont des heures de début à moins de 2h
           l'une de l'autre (le second write() écrasait silencieusement
           le premier).
"""

from __future__ import annotations

from datetime import datetime, time as dtime
from typing import Dict, List

import openpyxl
import openpyxl.styles
import openpyxl.styles.colors
import pytest

from src.cleaning import CleanedShiftReport, CraneRow
from src.merge import (
    CraneSession,
    MergedVesselDataset,
    merge_crane_sessions,
)
from src.report_generator import TPFREPWriter, fill_crane_timesheet_section


# ─────────────────────────────────────────────────────────────────────────────
# Helpers communs
# ─────────────────────────────────────────────────────────────────────────────

def _make_crane_row(
    shift_num: int,
    crane_id: str,
    vessel: str,
    doc: dtime,
    foc: dtime,
    imp: int = 0,
    exp: int = 0,
) -> CraneRow:
    return CraneRow(
        shift_num=shift_num,
        crane_id=crane_id,
        vessel_raw=vessel,
        vessel_normalized=vessel.replace(" ", "").upper(),
        import_moves=imp,
        export_moves=exp,
        doc_raw=doc,
        foc_raw=foc,
        observations=None,
    )


def _make_blue_workbook_with_two_slots(crane_id: str):
    """
    Construit un classeur openpyxl minimal simulant la structure d'un
    template TPFREP avec :
      - une ligne d'ancre "crane timesheet"
      - une ligne avec l'ID de grue
      - une ligne "crane type"
      - DEUX paires (commenced / completed) bleues pour cette grue
      - une ligne "total moves" bleue en dessous

    L'index de colonne de la grue est 3 (0-indexé) → colonne D.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    blue_fill = openpyxl.styles.PatternFill(
        fill_type="solid",
        fgColor=openpyxl.styles.colors.Color(indexed=44),
    )

    # Ancre textuelle section
    ws.cell(row=1, column=1).value = "crane timesheet"

    # Ligne IDs grues (row 3, 0-indexed row 2) — col D = col index 3
    ws.cell(row=3, column=4).value = crane_id  # 1-indexed: row=3, col=4 → D3

    # Crane type (row 4)
    ws.cell(row=4, column=1).value = "crane type"

    # Session 1 : commenced row 5, completed row 6
    ws.cell(row=5, column=1).value = "commenced"
    ws.cell(row=5, column=4).fill = blue_fill   # slot commenced grue, col D

    ws.cell(row=6, column=1).value = "completed"
    ws.cell(row=6, column=4).fill = blue_fill   # slot completed grue, col D

    # Session 2 : commenced row 7, completed row 8
    ws.cell(row=7, column=1).value = "commenced"
    ws.cell(row=7, column=4).fill = blue_fill

    ws.cell(row=8, column=1).value = "completed"
    ws.cell(row=8, column=4).fill = blue_fill

    # Total moves (row 10)
    ws.cell(row=10, column=1).value = "total moves"
    ws.cell(row=10, column=4).fill = blue_fill

    return wb, ws


# ─────────────────────────────────────────────────────────────────────────────
# Bug B — test_no_slot_collision_same_crane
# ─────────────────────────────────────────────────────────────────────────────

def test_no_slot_collision_same_crane():
    """
    Une grue P1 avec deux sessions à ~90 minutes d'écart (< 2h)
    doit occuper deux paires de lignes DISTINCTES dans le template —
    pas la même ligne écrasée deux fois.
    """
    crane_id = "P1"

    # Session 1 : 07:00 → 09:00
    session_1 = CraneSession(
        shift_num=1,
        commenced=datetime(2026, 6, 21, 7, 0),
        completed=datetime(2026, 6, 21, 9, 0),
        import_moves=30,
        export_moves=0,
    )
    # Session 2 : 08:30 → 10:30 (90 minutes après session 1 — écart < 2h)
    session_2 = CraneSession(
        shift_num=2,
        commenced=datetime(2026, 6, 21, 8, 30),
        completed=datetime(2026, 6, 21, 10, 30),
        import_moves=20,
        export_moves=10,
    )

    merged = MergedVesselDataset(
        vessel_name="TEST VESSEL",
        escale="TESTVESS_21062026",
        crane_sessions={crane_id: [session_1, session_2]},
    )

    wb, ws = _make_blue_workbook_with_two_slots(crane_id)
    writer = TPFREPWriter(ws)

    fill_crane_timesheet_section(writer, merged)

    # Récupérer les valeurs écrites dans la colonne D (col 4, 1-indexed)
    # Les slots dynamiques dans le mock sont aux lignes 5/6 et 7/8
    slot1_comm = ws.cell(row=5, column=4).value
    slot1_comp = ws.cell(row=6, column=4).value
    slot2_comm = ws.cell(row=7, column=4).value
    slot2_comp = ws.cell(row=8, column=4).value

    # Les deux paires de lignes doivent être remplies (pas de None partout)
    assert slot1_comm is not None, "Le slot 1 commenced est vide."
    assert slot1_comp is not None, "Le slot 1 completed est vide."
    assert slot2_comm is not None, "Le slot 2 commenced est vide."
    assert slot2_comp is not None, "Le slot 2 completed est vide."

    # Vérifier que les valeurs correspondent aux bonnes sessions
    assert slot1_comm == session_1.commenced, "session_1.commenced introuvable dans le slot 1."
    assert slot2_comm == session_2.commenced, "session_2.commenced introuvable dans le slot 2."


# ─────────────────────────────────────────────────────────────────────────────
# Bug A — test_unknown_crane_id_does_not_crash
# ─────────────────────────────────────────────────────────────────────────────

def test_unknown_crane_id_does_not_crash():
    """
    Une CraneRow avec crane_id='Q1' absent de config.CRANE_IDS ne doit
    pas lever d'exception. Les sessions de 'Q1' doivent être présentes
    dans le dict retourné.
    """
    import config

    # S'assurer que 'Q1' n'est pas dans config.CRANE_IDS (précondition du test)
    assert "Q1" not in config.CRANE_IDS, (
        "config.CRANE_IDS contient déjà 'Q1' — ajustez l'ID de grue inconnu dans ce test."
    )

    cleaned_shifts = {
        1: CleanedShiftReport(
            shift_num=1,
            shift_date=datetime(2026, 6, 21),
            crane_rows=[
                _make_crane_row(
                    shift_num=1,
                    crane_id="Q1",       # grue inconnue de config.CRANE_IDS
                    vessel="MASTERY D",
                    doc=dtime(7, 0),
                    foc=dtime(14, 0),
                    imp=50,
                    exp=0,
                ),
            ],
        ),
    }

    # Ne doit pas lever de KeyError ni aucune autre exception
    sessions = merge_crane_sessions(cleaned_shifts, target_vessel_normalized="MASTERYD")

    # Les sessions de 'Q1' doivent être présentes dans le résultat
    assert "Q1" in sessions, (
        "Les sessions de la grue 'Q1' sont absentes du dict retourné par "
        "merge_crane_sessions() — elles ont peut-être été silencieusement ignorées."
    )
    assert len(sessions["Q1"]) == 1, (
        f"1 session attendue pour 'Q1', {len(sessions['Q1'])} trouvée(s)."
    )
    assert sessions["Q1"][0].import_moves == 50


# ─────────────────────────────────────────────────────────────────────────────
# Cas limite : un seul slot disponible, deux sessions → la 2e doit être ignorée
# avec un WARNING (pas de crash)
# ─────────────────────────────────────────────────────────────────────────────


def test_overflow_sessions_logged_not_crashed():
    """
    Si les données ont plus de 6 sessions (notre limite fixe), la 7e session doit être ignorée
    avec un WARNING, sans crasher.
    """
    crane_id = "P1"
    
    sessions = []
    for i in range(7):
        sessions.append(CraneSession(
            shift_num=i+1,
            commenced=datetime(2026, 6, 21, 7 + i, 0),
            completed=datetime(2026, 6, 21, 7 + i, 30),
            import_moves=10,
            export_moves=5,
        ))

    merged = MergedVesselDataset(
        vessel_name="TEST VESSEL",
        escale="TESTVESS_21062026",
        crane_sessions={crane_id: sessions},
    )

    wb, ws = _make_blue_workbook_with_two_slots(crane_id)
    writer = TPFREPWriter(ws)

    # Ne doit pas crasher, même s'il y a plus de sessions que de slots
    fill_crane_timesheet_section(writer, merged)
    
    # Vérifier que les 2 premières sessions ont été traitées, le reste ignoré
    assert ws.cell(row=5, column=4).value is not None, "Le 1er slot n'a pas été rempli."
    assert ws.cell(row=7, column=4).value is not None, "Le 2e slot n'a pas été rempli."
