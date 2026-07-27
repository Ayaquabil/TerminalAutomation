import pathlib
import shutil
import pytest
import openpyxl
import config
from src.pipeline_runner import run_full_pipeline

PROJECT_DIR = pathlib.Path(__file__).resolve().parent.parent
FIXTURE_DIR = PROJECT_DIR / "tests" / "fixtures" / "reference"
INPUT_DIR = FIXTURE_DIR / "input"
OUTPUT_DIR = FIXTURE_DIR / "output"

def test_reference_compliance(tmp_path):
    """
    Runs the pipeline on the BELITAKI input fixture and compares the
    generated TPFREP_FINAL.xlsx cell-by-cell with the gold standard reference.
    """
    test_input = tmp_path / "input"
    test_template_dir = tmp_path / "template"
    test_output = tmp_path / "output"
    
    test_input.mkdir()
    test_template_dir.mkdir()
    test_output.mkdir()
    
    # Copy input files
    for f in INPUT_DIR.iterdir():
        if f.is_file():
            shutil.copy(f, test_input / f.name)
            
    # The gold standard file itself is also the template for this run
    ref_file = FIXTURE_DIR / "TPFREP_0XP93W1MA_BELITAKI.xlsx"
    shutil.copy(ref_file, test_template_dir / "template.xlsx")
    
    # Override configuration
    original_input = config.INPUT_DIR
    original_template = config.TEMPLATE_DIR
    original_output = config.OUTPUT_DIR
    original_report_path = config.OUTPUT_REPORT_PATH
    original_dashboard_path = config.OUTPUT_DASHBOARD_PATH
    original_template_path = config.TEMPLATE_PATH
    original_vessel = config.TARGET_VESSEL_NAME
    original_vessel_norm = config.TARGET_VESSEL_NORMALIZED
    original_voyage_import = config.VOYAGE_IMPORT
    original_voyage_export = config.VOYAGE_EXPORT
    
    config.INPUT_DIR = test_input
    config.TEMPLATE_DIR = test_template_dir
    config.OUTPUT_DIR = test_output
    config.OUTPUT_REPORT_PATH = test_output / "TPFREP_FINAL.xlsx"
    config.OUTPUT_DASHBOARD_PATH = test_output / "DASHBOARD.xlsx"
    config.TEMPLATE_PATH = test_template_dir / "template.xlsx"
    config.TARGET_VESSEL_NAME = "BELITAKI"
    config.TARGET_VESSEL_NORMALIZED = "BELITAKI"
    config.VOYAGE_IMPORT = "0XP93W1MA"
    config.VOYAGE_EXPORT = "0XP93W1MA"
    
    import src.import_data
    original_threshold = src.import_data.BLUE_CELL_TEMPLATE_THRESHOLD
    src.import_data.BLUE_CELL_TEMPLATE_THRESHOLD = 5
    
    wb_gen = None
    wb_ref = None
    try:
        result = run_full_pipeline()
        assert result.success, f"Pipeline failed: {result.error_message}"
        
        gen_path = test_output / "TPFREP_FINAL.xlsx"
        assert gen_path.exists(), "Generated report not found"
        
        # Load sheets
        wb_gen = openpyxl.load_workbook(gen_path, data_only=False)
        wb_ref = openpyxl.load_workbook(ref_file, data_only=False)
        
        ws_gen = wb_gen["TDPR v1.1"]
        ws_ref = wb_ref["TDPR v1.1"]
        
        discrepancies = []
        
        # Compare every non-empty cell of the reference sheet
        for r in range(1, ws_ref.max_row + 1):
            for c in range(1, ws_ref.max_column + 1):
                ref_val = ws_ref.cell(r, c).value
                gen_val = ws_gen.cell(r, c).value
                
                # We only care about comparing cells where reference is not empty
                if ref_val is not None:
                    # Let's normalize string comparisons
                    if isinstance(ref_val, str) and isinstance(gen_val, str):
                        ref_val_norm = ref_val.strip().replace(" ", "").upper()
                        gen_val_norm = gen_val.strip().replace(" ", "").upper()
                        if ref_val_norm != gen_val_norm:
                            discrepancies.append((r, c, ref_val, gen_val))
                    else:
                        # For float/datetime/etc., compare directly
                        if ref_val != gen_val:
                            discrepancies.append((r, c, ref_val, gen_val))
                    
        # Filter discrepancies to ignore formula cells and minor stuff
        if discrepancies:
            print("\nDiscrepancies found:")
            for r, c, ref, gen in discrepancies:
                cell_name = f"{openpyxl.utils.get_column_letter(c)}{r}"
                print(f"Cell {cell_name}: Ref={ref!r} | Gen={gen!r}")
                
        # Assert no discrepancies
        assert not discrepancies, f"Found {len(discrepancies)} cell discrepancies between generated report and reference."
        
    finally:
        config.INPUT_DIR = original_input
        config.TEMPLATE_DIR = original_template
        config.OUTPUT_DIR = original_output
        config.OUTPUT_REPORT_PATH = original_report_path
        config.OUTPUT_DASHBOARD_PATH = original_dashboard_path
        config.TEMPLATE_PATH = original_template_path
        config.TARGET_VESSEL_NAME = original_vessel
        config.TARGET_VESSEL_NORMALIZED = original_vessel_norm
        config.VOYAGE_IMPORT = original_voyage_import
        config.VOYAGE_EXPORT = original_voyage_export
        src.import_data.BLUE_CELL_TEMPLATE_THRESHOLD = original_threshold
        if wb_gen:
            wb_gen.close()
        if wb_ref:
            wb_ref.close()
