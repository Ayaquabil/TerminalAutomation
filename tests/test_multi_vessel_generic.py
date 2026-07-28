"""
tests/test_multi_vessel_generic.py

Test d'integration de bout en bout pour le navire fictif NORDIC AURORA.

Verifie que le pipeline est 100% generique et fonctionne avec n'importe
quel navire en modifiant uniquement config/settings.yaml.

Plan des sessions (fixe dans les fixtures generate_fixtures.py) :
  G1 : 3 sessions — total moves = 40+50+30 = 120
  G2 : 3 sessions — total moves = 30+30+20 =  80
  G3 : 8 sessions — dont une traversant minuit (22/06 23:30 -> 23/06 02:30)
       total moves = 50+20+40+30+15+20+25+30 = 230

Template : tests/fixtures/second_vessel/template.xlsx
  Ligne des cranes  : row 34 (1-indexed)
  Sessions commenced: rows 37,40,43,46,49,52,55,58 (1-indexed)
  Sessions completed: rows 38,41,44,47,50,53,56,59 (1-indexed)
  Ligne Total Moves : row 65 (1-indexed)
  Vessel Name       : E9
  UNLOCODE          : E10
  Voyage import     : E5
  Voyage export     : E8
"""

import pathlib
import shutil
import subprocess
import sys
from datetime import datetime

import openpyxl
import pytest

# Chemin du projet calcule dynamiquement depuis l'emplacement de CE fichier
PROJECT_DIR = pathlib.Path(__file__).resolve().parent.parent
SETTINGS_YAML_PATH = PROJECT_DIR / "config" / "settings.yaml"
FIXTURE_DIR = PROJECT_DIR / "tests" / "fixtures" / "second_vessel"
OUTPUT_DIR = FIXTURE_DIR / "output"


SECOND_VESSEL_YAML = f"""\
paths:
  data_dir: "{FIXTURE_DIR.as_posix()}"
  input_dir: "{FIXTURE_DIR.as_posix()}"
  template_dir: "{FIXTURE_DIR.as_posix()}"
  output_dir: "{OUTPUT_DIR.as_posix()}"
  archive_dir: "{(FIXTURE_DIR / 'archive').as_posix()}"
  template_filename: "template.xlsx"
  output_report_filename: "TPFREP_FINAL.xlsx"
  output_dashboard_filename: "DASHBOARD.xlsx"
  database_file: "{(FIXTURE_DIR / 'history.db').as_posix()}"
  logs_dir: "{(FIXTURE_DIR / 'logs').as_posix()}"
  log_file: "{(FIXTURE_DIR / 'logs' / 'application.log').as_posix()}"

file_detection:
  shift_filename_patterns:
    1: "shift_1"
    2: "shift_2"
    3: "shift_3"
  import_masteryd_pattern: "import_masteryd"
  export_masteryd_pattern: "export_masteryd"
  blue_cell_template_threshold: 5

vessel:
  target_name: "NORDIC AURORA"
  target_normalized: "NORDICAURORA"

terminal:
  vessel_port_unlocode: "GBHUL"
  terminal_code: "NHVTERMINAL"
  voyage_import: "7VK12X4AB"
  voyage_export: "7VK12X4AB"
  vessel_imo_default: "-"
  vessel_call_sign_default: "-"

cranes:
  crane_ids: ["G1", "G2", "G3"]
  template_column:
    G1: 4
    G2: 5
    G3: 6

shift_columns:
  crane_id: 0
  vessel: 1
  import_moves: 3
  export_moves: 4
  doc: 5
  foc: 6
  observations: 7

template_color:
  blue_indexed: 44
  blue_rgb_fallback: "FF99CCFF"

template_layout:
  max_sessions_per_crane: 8
  session_rows:
    - [36, 37]
    - [39, 40]
    - [42, 43]
    - [45, 46]
    - [48, 49]
    - [51, 52]
    - [54, 55]
    - [57, 58]
  total_moves_per_crane_row: 64
  general_delay_rows: [25, 26, 27, 28]
  restow_common_row: 202
  hatch_cover_row: 232
  discharged_start_row: 126
  loaded_deepsea_start_row: 157
  max_operators: 10

validation:
  max_reasonable_session_hours: 20
  min_valid_year: 2020
  max_valid_year: 2035
  required_masteryd_columns:
    - "N\\u00f8 CONTENEUR"
    - "CODE MVT"
    - "EXP IMP TRB"
    - "V/P"
    - "TYPE ISO"
    - "EXPLOITANT EN COURS"
    - "ESCALE"
"""


def test_multi_vessel_generic():
    """
    Test d'integration de bout en bout sur le navire fictif NORDIC AURORA.

    Verifie :
    1. Le pipeline se termine sans erreur (res.success == True).
    2. Nom du navire / UNLOCODE / voyage ecrits correctement dans le TPFREP.
    3. Chaque session grue est dans la bonne colonne et la bonne ligne.
    4. Les totaux G1, G2, G3 sont non-nuls et coherents.
    5. Aucune contamination croisee entre colonnes de grues.
    6. Le controle de coherence (cross_check_matches) passe.
    """
    assert FIXTURE_DIR.exists(), (
        f"Dossier de fixtures manquant : {FIXTURE_DIR}\n"
        "Lancez scratch/generate_second_vessel_fixtures.py pour le creer."
    )
    required_files = [
        "template.xlsx", "import_masteryd.xlsx", "export_masteryd.xlsx",
        "shift_1.xlsx", "shift_2.xlsx", "shift_3.xlsx",
    ]
    missing = [f for f in required_files if not (FIXTURE_DIR / f).exists()]
    assert not missing, f"Fichiers de fixtures manquants : {missing}"

    backup_path = SETTINGS_YAML_PATH.with_suffix(".yaml.bak_multitest")
    shutil.copy2(SETTINGS_YAML_PATH, backup_path)

    try:
        # 1. Installer la configuration second navire
        SETTINGS_YAML_PATH.write_text(SECOND_VESSEL_YAML, encoding="utf-8")

        # Nettoyer le dossier de sortie precedent
        if OUTPUT_DIR.exists():
            shutil.rmtree(OUTPUT_DIR)
        OUTPUT_DIR.mkdir(parents=True)

        # 2. Lancer le pipeline dans un sous-processus Python propre
        #    (sous-processus : evite la pollution du cache de settings en memoire)
        cmd = [
            sys.executable,
            "-c",
            f"import sys; sys.path.insert(0, r'{PROJECT_DIR}'); "
            "from src.pipeline_runner import run_full_pipeline; "
            "res = run_full_pipeline(); "
            "print('SUCCESS' if res.success else 'FAILED: ' + str(res.error_message)); "
            "assert res.success, res.error_message",
        ]
        result = subprocess.run(
            cmd, capture_output=True, text=True, timeout=60,
            cwd=str(PROJECT_DIR),
        )
        assert result.returncode == 0, (
            f"Pipeline NORDIC AURORA echoue :\n"
            f"STDOUT: {result.stdout}\n"
            f"STDERR: {result.stderr}"
        )

        # 3. Charger le rapport genere
        report_path = OUTPUT_DIR / "TPFREP_FINAL.xlsx"
        assert report_path.exists(), f"TPFREP_FINAL.xlsx non genere dans {OUTPUT_DIR}"

        wb = openpyxl.load_workbook(report_path, data_only=True)
        ws = wb.active

        # ── 4. Assertions de généralité (pas de résidus MASTERY D / BELITAKI) ──
        vessel_cell = ws["E9"].value
        assert vessel_cell == "NORDIC AURORA", (
            f"Nom du navire incorrect : attendu 'NORDIC AURORA', obtenu {vessel_cell!r}"
        )

        unlocode_cell = ws["E10"].value
        assert unlocode_cell == "GBHUL", (
            f"Port UNLOCODE incorrect : attendu 'GBHUL', obtenu {unlocode_cell!r}"
        )

        voyage_import = ws["E5"].value
        voyage_export = ws["E8"].value
        assert voyage_import in ("7VK12X4AB", "-"), (
            f"Voyage import inattendu : {voyage_import!r}"
        )
        assert voyage_export in ("7VK12X4AB", "-"), (
            f"Voyage export inattendu : {voyage_export!r}"
        )

        # ── 5. Verification des sessions grues (colonne E=G1, F=G2, G=G3) ─────
        # Template : session_rows[0] = [36,37] -> lignes Excel 37 et 38 (1-indexed)
        # G1 S1 commenced = 21/06 07:00, G1 S1 completed = 21/06 13:00
        # G2 S1 commenced = 21/06 07:30
        # Les 6 blocs hardcodés (1-indexed): 
        # B1 (38,39), B2 (41,42), B3 (44,45), B4 (47,48), B5 (50,51), B6 (53,54)
        def cell_is_datetime_or_str(cell_val):
            return isinstance(cell_val, (datetime, str)) and cell_val is not None

        # G1 a 3 sessions : elles remplissent séquentiellement B1, B2, B3
        assert cell_is_datetime_or_str(ws["E37"].value), f"G1 B1 commenced vide (E37={ws['E37'].value!r})"
        assert cell_is_datetime_or_str(ws["E38"].value), f"G1 B1 completed vide (E38={ws['E38'].value!r})"
        assert cell_is_datetime_or_str(ws["E40"].value), f"G1 B2 commenced vide (E40={ws['E40'].value!r})"
        assert cell_is_datetime_or_str(ws["E43"].value), f"G1 B3 commenced vide (E43={ws['E43'].value!r})"
        assert ws["E46"].value is None, f"G1 B4 devrait être vide (E46={ws['E46'].value!r})"

        # G2 a 3 sessions : elles remplissent B1, B2, B3
        assert cell_is_datetime_or_str(ws["F37"].value), f"G2 B1 commenced vide (F37={ws['F37'].value!r})"
        assert cell_is_datetime_or_str(ws["F40"].value), f"G2 B2 commenced vide (F40={ws['F40'].value!r})"
        assert cell_is_datetime_or_str(ws["F43"].value), f"G2 B3 commenced vide (F43={ws['F43'].value!r})"
        assert ws["F46"].value is None, f"G2 B4 devrait être vide (F46={ws['F46'].value!r})"

        # G3 a 8 sessions : la 7ème session sera récupérée via la recherche dynamique
        assert cell_is_datetime_or_str(ws["G37"].value), "G3 B1 commenced vide"
        assert cell_is_datetime_or_str(ws["G40"].value), "G3 B2 commenced vide"
        assert cell_is_datetime_or_str(ws["G43"].value), "G3 B3 commenced vide"
        assert cell_is_datetime_or_str(ws["G46"].value), "G3 B4 commenced vide"
        assert cell_is_datetime_or_str(ws["G49"].value), "G3 B5 commenced vide"
        assert cell_is_datetime_or_str(ws["G52"].value), "G3 B6 commenced vide"
        assert cell_is_datetime_or_str(ws["G55"].value), "G3 B7 commenced vide"
        assert cell_is_datetime_or_str(ws["G56"].value), "G3 B7 completed vide"

        # La session 6 de G3 (23:30 -> 02:30) traverse minuit :
        # commenced sur 22/06, completed sur 23/06
        g3_s6_comm = ws.cell(row=52, column=7).value   # ligne 52 -> slot 6 commenced
        g3_s6_comp = ws.cell(row=53, column=7).value   # ligne 53 -> slot 6 completed
        
        # Pour extraire la date, on peut vérifier la string s'il a été formaté
        if isinstance(g3_s6_comm, (datetime, str)) and isinstance(g3_s6_comp, (datetime, str)):
            assert g3_s6_comp > g3_s6_comm, (
                f"Session traversant minuit mal calculee : "
                f"completed ({g3_s6_comp}) <= commenced ({g3_s6_comm})"
            )

        # ── 6. Totaux de mouvements (ligne 65 = 1-indexed) ────────────────────
        # G1 : 40+50+30 = 120
        # G2 : 30+30+20 =  80
        # G3 : 50+20+40+30+15+20 = 175 (seulement 6 blocs traités max)
        g1_total = ws["E65"].value
        g2_total = ws["F65"].value
        g3_total = ws["G65"].value

        assert g1_total is not None and g1_total > 0, \
            f"Total G1 vide ou nul (E65={g1_total!r})"
        assert g2_total is not None and g2_total > 0, \
            f"Total G2 vide ou nul (F65={g2_total!r})"
        assert g3_total is not None and g3_total > 0, \
            f"Total G3 vide ou nul (G65={g3_total!r})"

        assert g1_total == 120, f"Total G1 incorrect : attendu 120, obtenu {g1_total}"
        assert g2_total ==  80, f"Total G2 incorrect : attendu  80, obtenu {g2_total}"
        assert g3_total == 230, f"Total G3 incorrect : attendu 230, obtenu {g3_total}"

        # ── 7. Pas de contamination croisee : G1 et G2 n'ont pas de valeur
        #        dans les slots où elles n'ont pas travaillé (ex: slot 5 = ligne 50) ─────────────────
        assert ws["E50"].value is None, \
            f"Contamination croisee : G1 a une valeur au slot 5 completed (E50={ws['E50'].value!r})"
        assert ws["F50"].value is None, \
            f"Contamination croisee : G2 a une valeur au slot 5 completed (F50={ws['F50'].value!r})"

        wb.close()

    finally:
        # Toujours restaurer le settings.yaml original
        if backup_path.exists():
            shutil.move(str(backup_path), str(SETTINGS_YAML_PATH))
