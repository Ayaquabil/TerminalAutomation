"""tests/test_report_generator.py — Tests unitaires du générateur de rapport TPFREP."""

import openpyxl
import pytest

from src.report_generator import TPFREPWriter, get_blue_cells, is_blue_cell


@pytest.fixture
def sample_workbook():
    """Classeur minimal en mémoire avec une cellule bleue (indexed=44), une
    cellule non bleue, et une cellule bleue contenant déjà une formule."""
    wb = openpyxl.Workbook()
    ws = wb.active

    blue_fill = openpyxl.styles.PatternFill(
        fill_type="solid", fgColor=openpyxl.styles.colors.Color(indexed=44)
    )
    other_fill = openpyxl.styles.PatternFill(
        fill_type="solid", fgColor=openpyxl.styles.colors.Color(indexed=9)
    )

    ws["A1"].fill = blue_fill          # cellule bleue vide -> doit être écrite
    ws["B1"].fill = other_fill         # cellule non bleue -> ne doit jamais être écrite
    ws["C1"].fill = blue_fill
    ws["C1"].value = "=SUM(A1:A2)"     # cellule bleue avec formule -> ne doit jamais être écrasée

    return wb, ws


class TestIsBlueCell:
    def test_detects_blue_fill(self, sample_workbook):
        _, ws = sample_workbook
        assert is_blue_cell(ws["A1"]) is True

    def test_rejects_other_fill(self, sample_workbook):
        _, ws = sample_workbook
        assert is_blue_cell(ws["B1"]) is False

    def test_rejects_no_fill(self, sample_workbook):
        _, ws = sample_workbook
        assert is_blue_cell(ws["D1"]) is False


class TestGetBlueCells:
    def test_collects_all_blue_coordinates(self, sample_workbook):
        _, ws = sample_workbook
        blue = get_blue_cells(ws)
        # A1 -> (0,0) 0-indexed ; C1 -> (0,2) 0-indexed
        assert (0, 0) in blue
        assert (0, 2) in blue
        assert (0, 1) not in blue  # B1 n'est pas bleue


class TestTPFREPWriter:
    def test_writes_to_blue_empty_cell(self, sample_workbook):
        _, ws = sample_workbook
        writer = TPFREPWriter(ws)
        result = writer.write(0, 0, "HELLO", "test")  # A1
        assert result is True
        assert ws["A1"].value == "HELLO"
        assert writer.filled_count == 1

    def test_refuses_to_write_to_non_blue_cell(self, sample_workbook):
        _, ws = sample_workbook
        writer = TPFREPWriter(ws)
        result = writer.write(0, 1, "SHOULD NOT WRITE", "test")  # B1
        assert result is False
        assert ws["B1"].value is None
        assert writer.skipped_count == 1

    def test_refuses_to_overwrite_existing_formula(self, sample_workbook):
        _, ws = sample_workbook
        writer = TPFREPWriter(ws)
        result = writer.write(0, 2, 999, "test")  # C1, bleue mais contient une formule
        assert result is False
        assert ws["C1"].value == "=SUM(A1:A2)"
        assert writer.skipped_count == 1

    def test_none_value_is_noop(self, sample_workbook):
        _, ws = sample_workbook
        writer = TPFREPWriter(ws)
        result = writer.write(0, 0, None, "test")
        assert result is False
        assert writer.filled_count == 0
