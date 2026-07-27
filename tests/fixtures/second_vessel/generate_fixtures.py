import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from pathlib import Path
from datetime import time, datetime

def create_fixtures():
    dest_dir = Path("C:/Users/ayaqu/Desktop/testbeltaki/TerminalAutomation/tests/fixtures/second_vessel")
    dest_dir.mkdir(parents=True, exist_ok=True)

    blue_fill = PatternFill(fill_type="solid", fgColor=Color(indexed=44))

    # ─────────────────────────────────────────────────────────────
    # 1. TEMPLATE.XLSX
    # ─────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TDPR v1.1"

    # En-têtes du template
    ws["C2"] = "TERMINAL DEPARTURE AND PERFORMANCE REPORT"
    ws["C3"] = "This spreadsheet is an exact mirror of the UN EDIFACT message TPFREP D 00B SMDG 3.0"
    
    # Identification
    ws["C5"] = "IMO No"
    ws["E5"].fill = blue_fill
    ws["C6"] = "Voyage # import"
    ws["E6"].fill = blue_fill
    ws["C7"] = "Call Sign"
    ws["E7"].fill = blue_fill
    ws["C8"] = "Voyage # export"
    ws["E8"].fill = blue_fill
    ws["C9"] = "Vessel Name"
    ws["E9"].fill = blue_fill
    ws["C10"] = "Port (UN Location Code)"
    ws["E10"].fill = blue_fill
    ws["C11"] = "Terminal Code"
    ws["E11"].fill = blue_fill
    ws["C12"] = "Vessel Operator"
    ws["E12"].fill = blue_fill

    # Crane Timesheet
    ws["C33"] = "2.1 Crane Timesheet"
    ws["C35"] = "Crane ID"
    ws["E35"] = "G1"
    ws["F35"] = "G2"
    ws["G35"] = "G3"

    ws["C36"] = "Crane Type"
    ws["E36"].fill = blue_fill
    ws["F36"].fill = blue_fill
    ws["G36"].fill = blue_fill

    # Sessions slots
    session_rows = [
        [37, 38], [40, 41], [43, 44], [46, 47],
        [49, 50], [52, 53], [55, 56], [58, 59]
    ]
    for pair in session_rows:
        r_start, r_end = pair
        # Col A-C labels
        ws.cell(row=r_start+1, column=3, value="commenced")
        ws.cell(row=r_end+1, column=3, value="completed")
        
        # Col E, F, G bleues
        for c in (5, 6, 7):
            ws.cell(row=r_start+1, column=c).fill = blue_fill
            ws.cell(row=r_end+1, column=c).fill = blue_fill

    # Total moves row
    ws["C65"] = "Total Moves per Crane"
    for c in (5, 6, 7):
        ws.cell(row=65, column=c).fill = blue_fill

    # Discharged Section
    ws["C127"] = "containers discharged"
    ws["C128"] = "discharged"
    ws["C129"] = "op"
    ws["D129"] = "full 20'"
    ws["E129"] = "full 40'"
    ws["F129"] = "empty 20'"
    ws["G129"] = "empty 40'"
    # Lignes bleues de données discharged
    for r in range(130, 133):
        for c in (3, 4, 5, 6, 7):
            ws.cell(row=r, column=c).fill = blue_fill

    # Loaded Section
    ws["C158"] = "containers loaded"
    ws["C159"] = "loaded"
    ws["C160"] = "op"
    ws["D160"] = "full 20'"
    ws["E160"] = "full 40'"
    ws["F160"] = "empty 20'"
    ws["G160"] = "empty 40'"
    # Lignes bleues de données loaded
    for r in range(161, 164):
        for c in (3, 4, 5, 6, 7):
            ws.cell(row=r, column=c).fill = blue_fill

    wb.save(dest_dir / "template.xlsx")
    wb.close()

    # ─────────────────────────────────────────────────────────────
    # 2. SHIFTS
    # ─────────────────────────────────────────────────────────────
    # Shift 1
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shift 1"
    ws["A1"] = "Portiques"
    ws["B1"] = datetime(2026, 6, 21)
    ws["A2"] = "Portique"
    ws["B2"] = "Navire"
    ws["D2"] = "Import"
    ws["E2"] = "Export"
    ws["F2"] = "DOC"
    ws["G2"] = "FOC"
    ws["H2"] = "Observations"

    # Grue G3 active
    ws["A3"] = "G3"
    ws["B3"] = "OCEANIC STAR"
    ws["D3"] = 50
    ws["E3"] = 0
    ws["F3"] = time(8, 0)
    ws["G3"] = time(12, 0)
    wb.save(dest_dir / "shift_1.xlsx")
    wb.close()

    # Shift 2
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shift 2"
    ws["A1"] = "Portiques"
    ws["B1"] = datetime(2026, 6, 22)
    ws["A2"] = "Portique"
    ws["B2"] = "Navire"
    ws["D2"] = "Import"
    ws["E2"] = "Export"
    ws["F2"] = "DOC"
    ws["G2"] = "FOC"
    ws["H2"] = "Observations"

    # Grue G3 active
    ws["A3"] = "G3"
    ws["B3"] = "OCEANIC STAR"
    ws["D3"] = 0
    ws["E3"] = 50
    ws["F3"] = time(13, 0)
    ws["G3"] = time(17, 0)
    wb.save(dest_dir / "shift_2.xlsx")
    wb.close()

    # Shift 3 (Traversée de minuit et chevauchements)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shift 3"
    ws["A1"] = "Portiques"
    ws["B1"] = datetime(2026, 6, 22)
    ws["A2"] = "Portique"
    ws["B2"] = "Navire"
    ws["D2"] = "Import"
    ws["E2"] = "Export"
    ws["F2"] = "DOC"
    ws["G2"] = "FOC"
    ws["H2"] = "Observations"

    # G1: 22:00 -> 04:00 (traversée de minuit)
    ws["A3"] = "G1"
    ws["B3"] = "OCEANIC STAR"
    ws["D3"] = 40
    ws["E3"] = 0
    ws["F3"] = time(22, 0)
    ws["G3"] = time(4, 0)

    # G2: 23:00 -> 03:00 (chevauchement avec G1)
    ws["A4"] = "G2"
    ws["B4"] = "OCEANIC STAR"
    ws["D4"] = 0
    ws["E4"] = 30
    ws["F4"] = time(23, 0)
    ws["G4"] = time(3, 0)

    # G3: 23:30 -> 02:30 (chevauchement avec G1 et G2)
    ws["A5"] = "G3"
    ws["B5"] = "OCEANIC STAR"
    ws["D5"] = 20
    ws["E5"] = 10
    ws["F5"] = time(23, 30)
    ws["G5"] = time(2, 30)
    
    wb.save(dest_dir / "shift_3.xlsx")
    wb.close()

    # ─────────────────────────────────────────────────────────────
    # 3. MASTERYD FILES
    headers = [
        "Nø CONTENEUR", "CODE MVT", "EXP IMP TRB", "V/P", "TYPE ISO", "Nø SCELLE ARMATEUR",
        "TAG FRIGO", "TAG DANG 0/1", "TAG HG 0/1", "EXPLOITANT EN COURS", "ESCALE", "CODE PORT CHAR",
        "DATE DE SAISIE", "HEURE DE SAISIE", "POOL"
    ]

    # Import MASTERYD (110 conteneurs)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for c_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=c_idx, value=h)
    
    # Remplir 110 conteneurs import
    for r in range(2, 112):
        ws.cell(row=r, column=1, value=f"CONTI{r:04d}")
        ws.cell(row=r, column=2, value="DEBA")
        ws.cell(row=r, column=3, value="I")
        ws.cell(row=r, column=4, value="P" if r % 2 == 0 else "V")
        ws.cell(row=r, column=5, value="22G1" if r % 2 == 0 else "45G1")
        ws.cell(row=r, column=6, value="SEAL")
        ws.cell(row=r, column=7, value=0)
        ws.cell(row=r, column=8, value=0)
        ws.cell(row=r, column=9, value=0)
        ws.cell(row=r, column=10, value="CMA" if r % 2 == 0 else "TAR")
        ws.cell(row=r, column=11, value="OCEANIC_STAR_24062026")
        ws.cell(row=r, column=12, value="MACAS")
        ws.cell(row=r, column=13, value=20260621)
        ws.cell(row=r, column=14, value=120000)
        ws.cell(row=r, column=15, value="G1" if r % 2 == 0 else "G3")
    wb.save(dest_dir / "import_masteryd.xlsx")
    wb.close()

    # Export MASTERYD (90 conteneurs)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for c_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=c_idx, value=h)
    
    # Remplir 90 conteneurs export
    for r in range(2, 92):
        ws.cell(row=r, column=1, value=f"CONTE{r:04d}")
        ws.cell(row=r, column=2, value="EMBA")
        ws.cell(row=r, column=3, value="E")
        ws.cell(row=r, column=4, value="P" if r % 2 == 0 else "V")
        ws.cell(row=r, column=5, value="22G1" if r % 2 == 0 else "45G1")
        ws.cell(row=r, column=6, value="SEAL")
        ws.cell(row=r, column=7, value=0)
        ws.cell(row=r, column=8, value=0)
        ws.cell(row=r, column=9, value=0)
        ws.cell(row=r, column=10, value="CMA" if r % 2 == 0 else "TAR")
        ws.cell(row=r, column=11, value="OCEANIC_STAR_24062026")
        ws.cell(row=r, column=12, value="MACAS")
        ws.cell(row=r, column=13, value=20260622)
        ws.cell(row=r, column=14, value=150000)
        ws.cell(row=r, column=15, value="G2" if r % 2 == 0 else "G3")
    wb.save(dest_dir / "export_masteryd.xlsx")
    wb.close()

    print("All fixtures created successfully.")

if __name__ == '__main__':
    create_fixtures()
