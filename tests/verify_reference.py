import pathlib
import shutil
import sys
import openpyxl

PROJECT_DIR = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_DIR))

import config
from src.pipeline_runner import run_full_pipeline
import src.import_data

def run_verification():
    FIXTURE_DIR = PROJECT_DIR / "tests" / "fixtures" / "reference"
    INPUT_DIR = FIXTURE_DIR / "input"
    
    tmp_path = PROJECT_DIR / "scratch_test_verify"
    if tmp_path.exists():
        shutil.rmtree(tmp_path)
    tmp_path.mkdir(parents=True)
    
    test_input = tmp_path / "input"
    test_template_dir = tmp_path / "template"
    test_output = tmp_path / "output"
    
    test_input.mkdir()
    test_template_dir.mkdir()
    test_output.mkdir()
    
    for f in INPUT_DIR.iterdir():
        if f.is_file():
            shutil.copy(f, test_input / f.name)
            
    ref_file = FIXTURE_DIR / "TPFREP_0XP93W1MA_BELITAKI.xlsx"
    shutil.copy(ref_file, test_template_dir / "template.xlsx")
    
    original_input = config.INPUT_DIR
    original_template = config.TEMPLATE_DIR
    original_output = config.OUTPUT_DIR
    original_report_path = config.OUTPUT_REPORT_PATH
    original_dashboard_path = config.OUTPUT_DASHBOARD_PATH
    original_template_path = config.TEMPLATE_PATH
    original_vessel = config.TARGET_VESSEL_NAME
    original_vessel_norm = config.TARGET_VESSEL_NORMALIZED
    original_voyage_import = config.VOYAGE_IMPORT
    original_voyage_export = config.VOYAGE_EXPORT
    original_threshold = src.import_data.BLUE_CELL_TEMPLATE_THRESHOLD
    
    config.INPUT_DIR = test_input
    config.TEMPLATE_DIR = test_template_dir
    config.OUTPUT_DIR = test_output
    config.OUTPUT_REPORT_PATH = test_output / "TPFREP_FINAL.xlsx"
    config.OUTPUT_DASHBOARD_PATH = test_output / "DASHBOARD.xlsx"
    config.TEMPLATE_PATH = test_template_dir / "template.xlsx"
    config.TARGET_VESSEL_NAME = "BELITAKI"
    config.TARGET_VESSEL_NORMALIZED = "BELITAKI"
    config.VOYAGE_IMPORT = "0XP93W1MA"
    config.VOYAGE_EXPORT = "0XP93W1MA"
    src.import_data.BLUE_CELL_TEMPLATE_THRESHOLD = 5
    
    wb_gen = None
    wb_ref = None
    try:
        result = run_full_pipeline(archive_after_success=False)
        print("Pipeline result success:", result.success)
        if not result.success:
            print("Error message:", result.error_message)
            return False
            
        gen_path = test_output / "TPFREP_FINAL.xlsx"
        wb_gen = openpyxl.load_workbook(gen_path, data_only=False)
        wb_ref = openpyxl.load_workbook(ref_file, data_only=False)
        
        ws_gen = wb_gen["TDPR v1.1"]
        ws_ref = wb_ref["TDPR v1.1"]
        
        discrepancies = []
        
        for r in range(1, ws_ref.max_row + 1):
            for c in range(1, ws_ref.max_column + 1):
                ref_val = ws_ref.cell(r, c).value
                gen_val = ws_gen.cell(r, c).value
                
                if ref_val is not None:
                    if isinstance(ref_val, str) and isinstance(gen_val, str):
                        ref_val_norm = ref_val.strip().replace(" ", "").upper()
                        gen_val_norm = gen_val.strip().replace(" ", "").upper()
                        if ref_val_norm != gen_val_norm:
                            discrepancies.append((r, c, ref_val, gen_val))
                    else:
                        if ref_val != gen_val:
                            discrepancies.append((r, c, ref_val, gen_val))
                            
        print(f"Total cell discrepancies found: {len(discrepancies)}")
        for r, c, ref, gen in discrepancies[:50]:
            cell_name = f"{openpyxl.utils.get_column_letter(c)}{r}"
            print(f"Cell {cell_name} (R{r}C{c}): Ref={ref!r} | Gen={gen!r}")
            
        return len(discrepancies) == 0
    finally:
        config.INPUT_DIR = original_input
        config.TEMPLATE_DIR = original_template
        config.OUTPUT_DIR = original_output
        config.OUTPUT_REPORT_PATH = original_report_path
        config.OUTPUT_DASHBOARD_PATH = original_dashboard_path
        config.TEMPLATE_PATH = original_template_path
        config.TARGET_VESSEL_NAME = original_vessel
        config.TARGET_VESSEL_NORMALIZED = original_vessel_norm
        config.VOYAGE_IMPORT = original_voyage_import
        config.VOYAGE_EXPORT = original_voyage_export
        src.import_data.BLUE_CELL_TEMPLATE_THRESHOLD = original_threshold
        if wb_gen:
            wb_gen.close()
        if wb_ref:
            wb_ref.close()
        if tmp_path.exists():
            shutil.rmtree(tmp_path, ignore_errors=True)

if __name__ == "__main__":
    run_verification()
