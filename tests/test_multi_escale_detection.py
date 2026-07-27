"""
tests/test_multi_escale_detection.py
=====================================

Tests pour la détection multi-escale, le filtrage des rapports de shift,
et le scénario d'intégration BELITAKI (mélange de navires/dates).

Stratégie de test :
  - Les tests T1–T5 utilisent des fichiers .xlsx synthétiques créés en
    mémoire avec openpyxl (tmp_path pytest), sans aucune dépendance aux
    données réelles. Cela garantit la reproductibilité et l'absence de
    faux-positifs.
  - T6 (scénario BELITAKI) utilise une fixture synthétique reproduisant
    le problème réel : MASTERYD de l'escale du 24/06 + rapports de shift
    des 21-23/06 portant d'autres navires (CMA CGM HAMLET, CC JANE).
  - T7 (intégration bout-en-bout) lance run_full_pipeline() dans un
    sous-processus Python propre avec la fixture second_vessel, sur le
    modèle de test_multi_vessel_generic.py.
"""

from __future__ import annotations

import json
import pathlib
import shutil
import subprocess
import sys
from datetime import datetime, time as dtime

import openpyxl
import pytest

from src.import_data import (
    EscaleInfo,
    InputDiscoveryError,
    MultipleEscalesError,
    discover_input_files,
)

PROJECT_DIR = pathlib.Path(__file__).resolve().parent.parent


# ─────────────────────────────────────────────────────────────────────────────
# Helpers : construction de fichiers .xlsx minimalistes
# ─────────────────────────────────────────────────────────────────────────────

def _write_shift_xlsx(path: pathlib.Path, vessel: str, shift_date: datetime) -> None:
    """Crée un fichier Excel de rapport de shift minimaliste."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shift 1"
    # En-tête reconnu par looks_like_shift_report()
    ws["A1"] = "SOMAPORT"
    ws["A2"] = shift_date          # date reconnue par le filtre de date
    ws["A4"] = "Portiques"         # ancre de détection shift
    ws["B4"] = "Navire"
    ws["C4"] = "ISO"
    ws["D4"] = "Import"
    ws["E4"] = "Export"
    ws["F4"] = "DOC"
    ws["G4"] = "FOC"
    ws["H4"] = "Observations"
    ws["A5"] = None                # sous-en-tête
    ws["F5"] = "Début"
    ws["G5"] = "Fin"
    # Ligne de données
    ws["A6"] = "P1"
    ws["B6"] = vessel
    ws["D6"] = 10
    ws["E6"] = 0
    ws["F6"] = dtime(7, 0)
    ws["G6"] = dtime(14, 0)
    wb.save(path)


def _write_masteryd_xlsx(
    path: pathlib.Path,
    direction: str,           # "I" pour IMPORT, "E" pour EXPORT
    escale: str,
    entry_date: datetime,
    n_rows: int = 5,
) -> None:
    """Crée un fichier Excel MASTERYD minimaliste."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MASTERYD"
    headers = [
        "N", "Nø CONTENEUR", "CODE MVT", "EXP IMP TRB", "V/P",
        "TYPE ISO", "Nø SCELLE ARMATEUR", "TAG FRIGO", "TAG DANG 0/1",
        "TAG HG 0/1", "EXPLOITANT EN COURS", "ESCALE",
        "CODE PORT DECHA", "AVARIES RESERVES", "DATE DE SAISIE", "HEURE DE SAISIE",
    ]
    ws.append(headers)
    for i in range(n_rows):
        ws.append([
            i + 1,
            f"CONT{i:07d}",
            "DEBA" if direction == "I" else "EMBA",
            direction,
            "V",
            "22G0",
            None, 0, 0, 0,
            "AKN",
            escale,
            "MACAS",
            None,
            entry_date,
            70000,
        ])
    wb.save(path)


def _write_template_xlsx(path: pathlib.Path) -> None:
    """Crée un template TPFREP minimaliste reconnu par looks_like_template()."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "TERMINAL DEPARTURE AND PERFORMANCE REPORT"
    wb.save(path)


# ─────────────────────────────────────────────────────────────────────────────
# T1 — Une seule escale : comportement identique à avant
# ─────────────────────────────────────────────────────────────────────────────

def test_single_escale_unaffected(tmp_path):
    """
    Un seul groupe MASTERYD + shifts cohérents : discover_input_files()
    doit retourner normalement, sans sélection requise, sans exception.
    """
    input_dir = tmp_path / "input"
    template_dir = tmp_path / "template"
    input_dir.mkdir()
    template_dir.mkdir()

    escale = "MASTERYD_24062026"
    entry_date = datetime(2026, 6, 24)

    _write_masteryd_xlsx(input_dir / "import.xlsx", "I", escale, entry_date, n_rows=10)
    _write_masteryd_xlsx(input_dir / "export.xlsx", "E", escale, entry_date, n_rows=8)
    _write_shift_xlsx(input_dir / "shift1.xlsx", "MASTERY D", entry_date)
    _write_template_xlsx(template_dir / "template.xlsx")

    # Ne doit pas lever d'exception
    result = discover_input_files(input_dir=input_dir, template_dir=template_dir)

    assert "import_masteryd" in result
    assert "export_masteryd" in result
    assert "shift_1" in result
    assert result["_chosen_escale"] == escale
    assert result["_excluded_shifts"] == []


# ─────────────────────────────────────────────────────────────────────────────
# T2 — Plusieurs escales : MultipleEscalesError avec la liste des escales
# ─────────────────────────────────────────────────────────────────────────────

def test_multiple_escales_requires_selection(tmp_path):
    """
    Deux paires de MASTERYD avec des ESCALE différentes :
    MultipleEscalesError doit être levée avec la liste des deux escales,
    et le pipeline ne doit PAS démarrer automatiquement.
    """
    input_dir = tmp_path / "input"
    template_dir = tmp_path / "template"
    input_dir.mkdir()
    template_dir.mkdir()

    escale_a = "BELITAKI_24062026"
    escale_b = "NORDICAURORA_28062026"
    date_a = datetime(2026, 6, 24)
    date_b = datetime(2026, 6, 28)

    # Escale A — 10 + 8 = 18 conteneurs
    _write_masteryd_xlsx(input_dir / "import_a.xlsx", "I", escale_a, date_a, n_rows=10)
    _write_masteryd_xlsx(input_dir / "export_a.xlsx", "E", escale_a, date_a, n_rows=8)
    # Escale B — 6 + 4 = 10 conteneurs
    _write_masteryd_xlsx(input_dir / "import_b.xlsx", "I", escale_b, date_b, n_rows=6)
    _write_masteryd_xlsx(input_dir / "export_b.xlsx", "E", escale_b, date_b, n_rows=4)

    # Un shift par escale — requis pour passer le check "Aucun rapport de shift reconnu"
    # qui précède le groupement multi-escale dans discover_input_files().
    _write_shift_xlsx(input_dir / "shift_a.xlsx", "BELITAKI",     date_a)
    _write_shift_xlsx(input_dir / "shift_b.xlsx", "NORDIC AURORA", date_b)

    _write_template_xlsx(template_dir / "template.xlsx")


    with pytest.raises(MultipleEscalesError) as exc_info:
        discover_input_files(input_dir=input_dir, template_dir=template_dir)

    err = exc_info.value
    escale_names = {e.name for e in err.escales}
    assert escale_a in escale_names, f"Escale A absente : {escale_names}"
    assert escale_b in escale_names, f"Escale B absente : {escale_names}"
    assert len(err.escales) == 2

    # Vérifier que les métadonnées (dates, conteneurs) sont présentes
    for info in err.escales:
        assert isinstance(info, EscaleInfo)
        assert info.container_count > 0
        assert info.min_date is not None


# ─────────────────────────────────────────────────────────────────────────────
# T3 — Shift hors plage de dates de l'escale → exclu avec raison
# ─────────────────────────────────────────────────────────────────────────────

def test_shift_report_outside_escale_date_range_excluded(tmp_path):
    """
    Un rapport de shift daté du 20/06 alors que l'escale MASTERYD est
    du 24/06 (+ marge 1 jour) : le shift doit figurer dans excluded_shifts,
    pas dans les shifts traités.
    """
    input_dir = tmp_path / "input"
    template_dir = tmp_path / "template"
    input_dir.mkdir()
    template_dir.mkdir()

    escale = "MASTERYD_24062026"
    escale_date = datetime(2026, 6, 24)
    stale_date = datetime(2026, 6, 20)    # 4 jours avant le min MASTERYD

    # MASTERYD daté du 24/06
    _write_masteryd_xlsx(input_dir / "import.xlsx", "I", escale, escale_date, n_rows=10)
    _write_masteryd_xlsx(input_dir / "export.xlsx", "E", escale, escale_date, n_rows=8)

    # Shift valide (24/06)
    _write_shift_xlsx(input_dir / "shift_ok.xlsx", "MASTERY D", escale_date)
    # Shift périmé (20/06, hors marge)
    _write_shift_xlsx(input_dir / "shift_old.xlsx", "MASTERY D", stale_date)

    _write_template_xlsx(template_dir / "template.xlsx")

    result = discover_input_files(input_dir=input_dir, template_dir=template_dir)

    # Le shift périmé doit être dans excluded_shifts
    excluded_files = [e["file"] for e in result["_excluded_shifts"]]
    assert "shift_old.xlsx" in excluded_files, (
        f"shift_old.xlsx aurait dû être exclu (hors plage). "
        f"Exclus trouvés : {excluded_files}"
    )

    # Le shift valide doit être présent
    shift_keys = [k for k in result if k.startswith("shift_")]
    assert len(shift_keys) == 1, (
        f"Seul le shift valide devrait être traité, trouvé : {shift_keys}"
    )

    # La raison de l'exclusion doit mentionner la plage de dates
    old_exclusion = next(e for e in result["_excluded_shifts"] if e["file"] == "shift_old.xlsx")
    assert "plage" in old_exclusion["reason"].lower() or "date" in old_exclusion["reason"].lower(), (
        f"La raison d'exclusion devrait mentionner la plage/date : {old_exclusion['reason']!r}"
    )


# ─────────────────────────────────────────────────────────────────────────────
# T4 — Shift avec navire non correspondant → exclu et logué
# ─────────────────────────────────────────────────────────────────────────────

def test_shift_report_wrong_vessel_excluded_and_logged(tmp_path):
    """
    Un rapport de shift dans la bonne fenêtre de dates mais dont aucune
    ligne ne correspond au navire de l'escale → doit être dans excluded_shifts
    avec une raison explicite (pas juste silencieusement absent).
    """
    input_dir = tmp_path / "input"
    template_dir = tmp_path / "template"
    input_dir.mkdir()
    template_dir.mkdir()

    escale = "MASTERYD_24062026"
    escale_date = datetime(2026, 6, 24)

    _write_masteryd_xlsx(input_dir / "import.xlsx", "I", escale, escale_date, n_rows=10)
    _write_masteryd_xlsx(input_dir / "export.xlsx", "E", escale, escale_date, n_rows=8)

    # Shift valide (bon navire)
    _write_shift_xlsx(input_dir / "shift_correct.xlsx", "MASTERY D", escale_date)
    # Shift étranger : même date mais navire différent (CMA CGM HAMLET)
    _write_shift_xlsx(input_dir / "shift_foreign.xlsx", "CMA CGM HAMLET", escale_date)

    _write_template_xlsx(template_dir / "template.xlsx")

    result = discover_input_files(input_dir=input_dir, template_dir=template_dir)

    excluded_files = [e["file"] for e in result["_excluded_shifts"]]
    assert "shift_foreign.xlsx" in excluded_files, (
        f"shift_foreign.xlsx (CMA CGM HAMLET) aurait dû être exclu. "
        f"Exclus trouvés : {excluded_files}"
    )

    # La raison doit mentionner le navire
    foreign_exclusion = next(
        e for e in result["_excluded_shifts"] if e["file"] == "shift_foreign.xlsx"
    )
    reason = foreign_exclusion["reason"].lower()
    assert "navire" in reason or "vessel" in reason or "cible" in reason, (
        f"La raison d'exclusion devrait mentionner le navire : {foreign_exclusion['reason']!r}"
    )

    # Le shift correct doit être présent
    shift_keys = [k for k in result if k.startswith("shift_")]
    assert len(shift_keys) == 1


# ─────────────────────────────────────────────────────────────────────────────
# T5 — Vrai doublon de shift : InputDiscoveryError bloquante
# ─────────────────────────────────────────────────────────────────────────────

def test_real_duplicate_shift_number_raises(tmp_path):
    """
    Deux fichiers différents, même numéro de shift, même heure DOC
    (= même sort_key) → InputDiscoveryError bloquante, pas de renumérotage.
    """
    input_dir = tmp_path / "input"
    template_dir = tmp_path / "template"
    input_dir.mkdir()
    template_dir.mkdir()

    escale = "MASTERYD_24062026"
    escale_date = datetime(2026, 6, 24)

    _write_masteryd_xlsx(input_dir / "import.xlsx", "I", escale, escale_date, n_rows=10)
    _write_masteryd_xlsx(input_dir / "export.xlsx", "E", escale, escale_date, n_rows=8)
    _write_template_xlsx(template_dir / "template.xlsx")

    # Deux shifts avec le même numéro ET la même heure DOC → doublon strict
    def _write_identical_shift(path: pathlib.Path) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Shift 1"
        ws["A1"] = "SOMAPORT"
        ws["A2"] = escale_date
        ws["A3"] = "SHIFT 1"           # numéro de shift explicite
        ws["A4"] = "Portiques"
        ws["B4"] = "Navire"
        ws["F4"] = "DOC"
        ws["G4"] = "FOC"
        ws["A5"] = None
        ws["F5"] = "Début"
        ws["G5"] = "Fin"
        ws["A6"] = "P1"
        ws["B6"] = "MASTERY D"
        ws["D6"] = 10
        ws["F6"] = dtime(7, 0)         # même DOC = même sort_key
        ws["G6"] = dtime(14, 0)
        wb.save(path)

    _write_identical_shift(input_dir / "shift_1a.xlsx")
    _write_identical_shift(input_dir / "shift_1b.xlsx")

    with pytest.raises(InputDiscoveryError) as exc_info:
        discover_input_files(input_dir=input_dir, template_dir=template_dir)

    msg = str(exc_info.value).lower()
    assert "doublon" in msg or "conflit" in msg or "identique" in msg, (
        f"Le message d'erreur devrait mentionner le doublon : {exc_info.value!r}"
    )


# ─────────────────────────────────────────────────────────────────────────────
# T6 — Scénario BELITAKI synthétique : mélange de navires et de dates
# ─────────────────────────────────────────────────────────────────────────────

def test_belitaki_mixed_scenario(tmp_path):
    """
    Reproduit le problème réel observé sur BELITAKI avec une fixture
    entièrement synthétique (pas de dépendance aux données réelles) :

      - MASTERYD IMPORT/EXPORT de l'escale BELITAKI_24062026 (entré le 24/06)
      - Rapports de shift des 21, 22, 23/06 portant d'autres navires
        (CMA CGM HAMLET, CC JANE) présents dans le même dossier
      - Un rapport de shift du 24/06 portant BELITAKI (le bon)

    Attendu :
      - Seul le shift du 24/06 est retenu.
      - Les shifts des 21-23/06 sont dans excluded_shifts (hors plage de dates
        ou navire non correspondant).
      - La raison d'exclusion est renseignée pour chaque fichier exclu.
    """
    input_dir = tmp_path / "input"
    template_dir = tmp_path / "template"
    input_dir.mkdir()
    template_dir.mkdir()

    escale = "BELITAKI_24062026"
    escale_date = datetime(2026, 6, 24)

    # MASTERYD synthétiques de l'escale du 24/06
    _write_masteryd_xlsx(input_dir / "IMPORT_MASTERYD.xlsx", "I", escale, escale_date, n_rows=50)
    _write_masteryd_xlsx(input_dir / "EXPORT_MASTERYD.xlsx", "E", escale, escale_date, n_rows=10)

    _write_template_xlsx(template_dir / "template.xlsx")

    # Shifts "étrangers" des 21-23/06 (autres navires, hors plage)
    _write_shift_xlsx(input_dir / "shift_21jun_hamlet.xlsx",  "CMA CGM HAMLET", datetime(2026, 6, 21))
    _write_shift_xlsx(input_dir / "shift_22jun_ccjane.xlsx",  "CC JANE",         datetime(2026, 6, 22))
    _write_shift_xlsx(input_dir / "shift_23jun_hamlet.xlsx",  "CMA CGM HAMLET", datetime(2026, 6, 23))

    # Le bon shift (24/06, BELITAKI)
    _write_shift_xlsx(input_dir / "shift_24jun_belitaki.xlsx", "BELITAKI", escale_date)

    result = discover_input_files(input_dir=input_dir, template_dir=template_dir)

    excluded_files = {e["file"] for e in result["_excluded_shifts"]}
    shift_keys = [k for k in result if k.startswith("shift_")]

    # Les trois shifts étrangers/périmés doivent être exclus
    assert "shift_21jun_hamlet.xlsx"  in excluded_files, f"Non exclu : {excluded_files}"
    assert "shift_22jun_ccjane.xlsx"  in excluded_files, f"Non exclu : {excluded_files}"
    assert "shift_23jun_hamlet.xlsx"  in excluded_files, f"Non exclu : {excluded_files}"

    # Seul le shift BELITAKI du 24/06 est retenu
    assert len(shift_keys) == 1, (
        f"Un seul shift attendu (BELITAKI 24/06), trouvé : {shift_keys}"
    )

    # La raison d'exclusion est renseignée pour chaque exclu
    for excl in result["_excluded_shifts"]:
        assert excl["reason"], f"Raison vide pour {excl['file']}"


# ─────────────────────────────────────────────────────────────────────────────
# T7 — Intégration bout-en-bout : pipeline avec shift intrus, second_vessel
# ─────────────────────────────────────────────────────────────────────────────

FIXTURE_DIR = PROJECT_DIR / "tests" / "fixtures" / "second_vessel"

SECOND_VESSEL_YAML_TEMPLATE = """\
paths:
  data_dir: "{fixture_dir}"
  input_dir: "{fixture_dir}"
  template_dir: "{fixture_dir}"
  output_dir: "{output_dir}"
  archive_dir: "{archive_dir}"
  template_filename: "template.xlsx"
  output_report_filename: "TPFREP_FINAL.xlsx"
  output_dashboard_filename: "DASHBOARD.xlsx"
  database_file: "{fixture_dir}/history.db"
  logs_dir: "{fixture_dir}/logs"
  log_file: "{fixture_dir}/logs/application.log"

file_detection:
  shift_filename_patterns:
    1: "shift_1"
    2: "shift_2"
    3: "shift_3"
  import_masteryd_pattern: "import_masteryd"
  export_masteryd_pattern: "export_masteryd"
  blue_cell_template_threshold: 5

vessel:
  target_name: "NORDIC AURORA"
  target_normalized: "NORDICAURORA"

terminal:
  vessel_port_unlocode: "GBHUL"
  terminal_code: "NHVTERMINAL"
  voyage_import: "7VK12X4AB"
  voyage_export: "7VK12X4AB"
  vessel_imo_default: "-"
  vessel_call_sign_default: "-"

cranes:
  crane_ids: ["G1", "G2", "G3"]
  template_column:
    G1: 4
    G2: 5
    G3: 6

shift_columns:
  crane_id: 0
  vessel: 1
  import_moves: 3
  export_moves: 4
  doc: 5
  foc: 6
  observations: 7

template_color:
  blue_indexed: 44
  blue_rgb_fallback: "FF99CCFF"

template_layout:
  max_sessions_per_crane: 8
  session_rows:
    - [36, 37]
    - [39, 40]
    - [42, 43]
    - [45, 46]
    - [48, 49]
    - [51, 52]
    - [54, 55]
    - [57, 58]
  total_moves_per_crane_row: 64
  general_delay_rows: [25, 26, 27, 28]
  restow_common_row: 202
  hatch_cover_row: 232
  discharged_start_row: 126
  loaded_deepsea_start_row: 157
  max_operators: 10

validation:
  max_reasonable_session_hours: 20
  min_valid_year: 2020
  max_valid_year: 2035
  shift_date_margin_days: 1
  required_masteryd_columns:
    - "N\\u00f8 CONTENEUR"
    - "CODE MVT"
    - "EXP IMP TRB"
    - "V/P"
    - "TYPE ISO"
    - "EXPLOITANT EN COURS"
    - "ESCALE"
"""

SETTINGS_YAML_PATH = PROJECT_DIR / "config" / "settings.yaml"


@pytest.mark.skipif(
    not FIXTURE_DIR.exists()
    or not (FIXTURE_DIR / "template.xlsx").exists()
    or not (FIXTURE_DIR / "import_masteryd.xlsx").exists(),
    reason=(
        "Fixtures second_vessel absentes — lancez "
        "scratch/generate_second_vessel_fixtures.py pour les créer."
    ),
)
def test_end_to_end_pipeline_with_foreign_shift(tmp_path):
    """
    Test d'intégration bout-en-bout (point 1 demandé explicitement) :

    Lance run_full_pipeline() dans un sous-processus Python propre sur les
    fixtures second_vessel (NORDIC AURORA), en ajoutant UN rapport de shift
    étranger (navire fictif 'INTRUDER VESSEL', même plage de dates).

    Vérifie :
      a) Le pipeline se termine avec succès (res.success == True).
      b) Le shift étranger est dans result.excluded_shifts.
      c) Les totaux G1/G2/G3 sont inchangés (120 / 80 / 230), ce qui prouve
         que le shift étranger n'a pas contaminé les calculs.
    """
    work_dir = tmp_path / "work"
    work_dir.mkdir()
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    archive_dir = tmp_path / "archive"
    archive_dir.mkdir()
    (tmp_path / "logs").mkdir(exist_ok=True)

    # Copier toutes les fixtures dans work_dir
    required = [
        "template.xlsx", "import_masteryd.xlsx", "export_masteryd.xlsx",
        "shift_1.xlsx", "shift_2.xlsx", "shift_3.xlsx",
    ]
    for fname in required:
        shutil.copy(FIXTURE_DIR / fname, work_dir / fname)

    # Injecter un shift étranger (même plage de dates, mauvais navire)
    _write_shift_xlsx(
        work_dir / "shift_intruder.xlsx",
        vessel="INTRUDER VESSEL",
        shift_date=datetime(2026, 6, 21),
    )

    backup = SETTINGS_YAML_PATH.with_suffix(".yaml.bak_t7")
    shutil.copy2(SETTINGS_YAML_PATH, backup)

    yaml_content = SECOND_VESSEL_YAML_TEMPLATE.format(
        fixture_dir=work_dir.as_posix(),
        output_dir=output_dir.as_posix(),
        archive_dir=archive_dir.as_posix(),
    )

    try:
        SETTINGS_YAML_PATH.write_text(yaml_content, encoding="utf-8")

        script = (
            f"import sys; sys.path.insert(0, r'{PROJECT_DIR}'); "
            "from src.pipeline_runner import run_full_pipeline; "
            "res = run_full_pipeline(); "
            "import json; "
            "excluded = [e['file'] for e in (res.excluded_shifts or [])]; "
            "print('STATUS:' + ('SUCCESS' if res.success else 'FAILED')); "
            "print('EXCLUDED:' + json.dumps(excluded)); "
            "print('G1:' + str(res.kpi.cross_check_crane_moves_discharged if res.kpi else None)); "
            "assert res.success, res.error_message"
        )
        proc = subprocess.run(
            [sys.executable, "-c", script],
            capture_output=True, text=True, timeout=120,
            cwd=str(PROJECT_DIR),
        )
        assert proc.returncode == 0, (
            f"Pipeline T7 échoué :\nSTDOUT: {proc.stdout}\nSTDERR: {proc.stderr}"
        )

        lines = proc.stdout.strip().splitlines()
        status_line = next((l for l in lines if l.startswith("STATUS:")), "")
        excluded_line = next((l for l in lines if l.startswith("EXCLUDED:")), "")

        # a) Pipeline réussi
        assert "SUCCESS" in status_line, f"Pipeline non réussi : {proc.stdout}"

        # b) Shift intrus exclu
        excluded_files = json.loads(excluded_line[len("EXCLUDED:"):]) if excluded_line else []
        assert "shift_intruder.xlsx" in excluded_files, (
            f"shift_intruder.xlsx n'est pas dans excluded_shifts : {excluded_files}"
        )

        # c) Totaux G1/G2/G3 inchangés
        report_path = output_dir / "TPFREP_FINAL.xlsx"
        assert report_path.exists(), "TPFREP_FINAL.xlsx non généré"

        wb = openpyxl.load_workbook(report_path, data_only=True)
        ws = wb.active
        g1_total = ws["E65"].value
        g2_total = ws["F65"].value
        g3_total = ws["G65"].value
        wb.close()

        assert g1_total == 120, f"Total G1 faussé par le shift intrus : {g1_total}"
        assert g2_total == 80,  f"Total G2 faussé par le shift intrus : {g2_total}"
        assert g3_total == 230, f"Total G3 faussé par le shift intrus : {g3_total}"

    finally:
        if backup.exists():
            shutil.move(str(backup), str(SETTINGS_YAML_PATH))
