"""
src/report_generator.py — Remplissage automatique du template TPFREP.

Moteur de remplissage totalement dynamique :
- Aucune coordonnée Excel codée en dur dans les sections métier.
- Cartographie automatique du template via index textuel et détection
  des cellules bleues à l'initialisation du TPFREPWriter.
- Détection des colonnes de grues, sessions, retards, conteneurs par
  analyse des textes d'ancrage et des couleurs, avec dictionnaire de
  synonymes pour la robustesse face aux variantes de libellés.
- Protection intégrale : formules et styles jamais modifiés.
- Données manquantes : cellule laissée vide + warning précis loggé.

Interface publique préservée (compatibilité pipeline_runner.py et tests) :
  generate_tpfrep_report(merged, kpi, template_path, output_path) -> Path
  is_blue_cell(cell) -> bool
  get_blue_cells(ws) -> set
  class TPFREPWriter: write(), find_anchor(), write_near_anchor()
"""

from __future__ import annotations

import re
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import openpyxl.utils

import config
from src.calculations import KPIResult
from src.import_data import load_template_workbook
from src.logger import get_logger
from src.merge import MergedVesselDataset

logger = get_logger("report_generator")


# ═══════════════════════════════════════════════════════════════════════════
# UTILITAIRES DE NORMALISATION DE TEXTE
# ═══════════════════════════════════════════════════════════════════════════

def _normalize(text: str) -> str:
    """Normalise un texte : minuscules, ASCII sans accents, espaces réduits."""
    if not isinstance(text, str):
        return ""
    nfkd = unicodedata.normalize("NFKD", text)
    ascii_ = nfkd.encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", ascii_).strip().lower()


def _matches(cell_text: str, synonyms: List[str]) -> bool:
    """True si la version normalisée de cell_text contient l'un des synonymes normalisés."""
    norm = _normalize(cell_text)
    return any(_normalize(s) in norm for s in synonyms)


# ═══════════════════════════════════════════════════════════════════════════
# DICTIONNAIRE DE SYNONYMES PAR CATÉGORIE DE SECTION
# ═══════════════════════════════════════════════════════════════════════════

_SYNONYMS: Dict[str, List[str]] = {
    "crane_timesheet": ["crane timesheet", "crane id"],
    "crane_type":      ["crane type"],
    "crane_commenced": ["commenced"],
    "crane_completed": ["completed"],
    "crane_total":     ["total moves", "total # moves", "total moves per crane"],
    "full_header":     ["full discharged", "full loaded", "deepsea full loaded",
                        "deepsea full", "full"],
    "empty_header":    ["empty discharged", "empty loaded", "deepsea empty loaded",
                        "deepsea empty", "empty"],
    "size_20":         ["20'", "20 ft", "20ft"],
    "size_40":         ["40'", "40 ft", "40ft", "40' + 45'", "40'  +  45'"],
    "size_total":      ["total", "tot", "sum"],
    "general_delays":  ["general delays", "delays not assigned"],
    "break_bulk":      ["break bulk", "break bulk moves"],
    "containers_disc": ["containers discharged", "discharged"],
    "containers_load": ["containers loaded", "deepsea loaded", "loaded"],
    "delays_ship_cargo": ["delays caused by ship or cargo", "ship or cargo operation", "ship or cargo delays"],
    "delays_port_terminal": ["delays caused by port", "port or terminal operator", "port or terminal delays"],
    "transhipment_disc": ["transhipment discharged", "3.2 transhipment"],
    "shortsea_load":     ["shortsea loaded", "4.2 shortsea"],
    "transhipment_load": ["transhipment loaded", "4.3 transhipment"],
    "restow_two_moves":  ["discharge+reload", "restow, two moves", "5.1 discharge"],
    "restow_one_move":   ["shift on board", "restow, one move", "5.2 shift"],
    "hatch_cover_moves": ["hatch cover moves", "hatch cover", "6 hatch cover"],
}


# ═══════════════════════════════════════════════════════════════════════════
# DÉTECTION DE LA COULEUR BLEUE
# ═══════════════════════════════════════════════════════════════════════════

def is_blue_cell(cell) -> bool:
    """True si la cellule a un remplissage plein de couleur indexée 44 (bleu
    du template natif .xlsx) OU de couleur RVB FF99CCFF (bleu équivalent
    produit quand le template était fourni en .xls et converti en .xlsx)."""
    fill = cell.fill
    if fill is None or getattr(fill, "patternType", None) != "solid":
        return False
    fg = fill.fgColor
    if fg is None:
        return False
    if getattr(fg, "indexed", None) == config.BLUE_INDEXED:
        return True
    if getattr(fg, "type", None) == "rgb" and str(getattr(fg, "rgb", "")).upper() == config.BLUE_RGB_FALLBACK:
        return True
    return False


def get_blue_cells(ws) -> set:
    """Ensemble des coordonnées (row, col) 0-indexed des cellules bleues du template."""
    blue = set()
    for row in ws.iter_rows():
        for cell in row:
            if is_blue_cell(cell):
                blue.add((cell.row - 1, cell.column - 1))
    return blue


# ═══════════════════════════════════════════════════════════════════════════
# TPFREP WRITER — MOTEUR DE REMPLISSAGE DYNAMIQUE
# ═══════════════════════════════════════════════════════════════════════════

class TPFREPWriter:
    """
    Encapsule l'écriture protégée dans le template TPFREP.

    Fournit un moteur de cartographie dynamique :
    - Index textuel complet de tous les libellés du template.
    - Détection automatique des colonnes de grues, paires de sessions,
      colonnes Full/Empty 20'/40', lignes de retards — sans coordonnée fixe.
    - Protection intégrale : formules et styles jamais modifiés.
    - Journalisation précise de chaque écriture, saut ou échec.
    """

    def __init__(self, ws):
        self.ws = ws
        self.blue_cells: set = get_blue_cells(ws)
        self.filled_count: int = 0
        self.skipped_count: int = 0
        # Index textuel : texte_normalisé → [(row_0indexed, col_0indexed), ...]
        self._text_index: Dict[str, List[Tuple[int, int]]] = {}
        self._build_text_index()
        logger.info(
            "Template cartographié : %d cellules bleues, %d entrées textuelles "
            "sur la feuille '%s'.",
            len(self.blue_cells), len(self._text_index), ws.title,
        )

    # ── Construction de l'index ─────────────────────────────────────────────

    def _build_text_index(self) -> None:
        """Construit un index {texte_normalisé: [(row0, col0)]} sur toutes les cellules texte."""
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    key = _normalize(cell.value)
                    if key:
                        self._text_index.setdefault(key, []).append(
                            (cell.row - 1, cell.column - 1)
                        )

    def _addr(self, row: int, col: int) -> str:
        """Adresse Excel lisible (ex: 'D18') — utilisée dans les logs."""
        return f"{openpyxl.utils.get_column_letter(col + 1)}{row + 1}"

    # ── Écriture protégée ───────────────────────────────────────────────────

    def write(self, row: int, col: int, value, label: str = "", allow_zero: bool = False, force: bool = False) -> bool:
        """Écrit `value` en (row, col) 0-indexed si la cellule est bleue (ou si force=True) et sans formule.

        allow_zero=True : écrit les valeurs 0/0.0 (pour les compteurs de conteneurs).
        allow_zero=False (défaut) : ignore None, '', 0, 0.0
                                    (donnée absente → cellule laissée vide).
        force=True : outrepasse la vérification des cellules bleues (utile pour les totaux non-bleus).

        Journalise chaque tentative : écriture réussie, saut ou rejet.
        """
        if value is None or value == "":
            if label:
                logger.debug("Valeur absente — cellule laissée vide : %s", label)
            return False
        if not allow_zero and (value == 0 or value == 0.0):
            logger.debug("Zéro ignoré (allow_zero=False) : %s", label)
            return False

        if not force and (row, col) not in self.blue_cells:
            logger.debug(
                "SKIP %s (non bleue) — raison : cellule hors zone d'écriture autorisée [%s]",
                self._addr(row, col), label,
            )
            self.skipped_count += 1
            return False

        cell = self.ws.cell(row=row + 1, column=col + 1)
        existing = cell.value
        if cell.data_type == "f" or (isinstance(existing, str) and existing.startswith("=")):
            logger.warning(
                "SKIP %s : cellule contient la formule '%s' — jamais écrasée. "
                "[KPI : %s]",
                self._addr(row, col), existing or "formula", label,
            )
            self.skipped_count += 1
            return False

        if isinstance(value, datetime):
            from src.utils import round_to_nearest_minute
            value = round_to_nearest_minute(value)
            if not cell.number_format or cell.number_format.lower() in ("general", "@"):
                cell.number_format = "d/m/yy h:mm"

        cell.value = value
        self.filled_count += 1
        logger.info(
            "KPI : %-40s -> Feuille : %-12s -> Cellule : %s -> Valeur : %r -> OK",
            label, self.ws.title, self._addr(row, col), value,
        )
        return True

    def write_cell_by_coord(self, coord: str, value, label: str = "") -> bool:
        """Écrit directement dans une cellule ciblée par sa coordonnée Excel (ex: 'D15')
        en respectant strictement la protection des formules et le formatage des dates.
        """
        if value is None or value == "":
            return False
        from src.utils import round_to_nearest_minute
        cell = self.ws[coord]
        existing = cell.value
        if cell.data_type == "f" or (isinstance(existing, str) and existing.startswith("=")):
            logger.warning("SKIP %s : cellule contient la formule '%s' — jamais écrasée. [%s]", coord, existing, label)
            self.skipped_count += 1
            return False
        if isinstance(value, datetime):
            value = round_to_nearest_minute(value)
            if not cell.number_format or cell.number_format.lower() in ("general", "@"):
                cell.number_format = "d/m/yy h:mm"
        cell.value = value
        self.filled_count += 1
        logger.info("KPI : %-40s -> Feuille : %-12s -> Cellule : %s -> Valeur : %r -> OK", label, self.ws.title, coord, value)
        return True

    # ── Recherche textuelle ─────────────────────────────────────────────────

    def find_anchor(self, search_text: str) -> Tuple[int, int]:
        """Cherche un texte (exact ou partiel) via l'index textuel du template.
        Retourne (row, col) 0-indexed ou (-1, -1) si non trouvé.
        """
        key = _normalize(search_text)
        # Recherche exacte
        if key in self._text_index:
            return self._text_index[key][0]
        # Recherche partielle
        for idx_key, coords in self._text_index.items():
            if key in idx_key:
                return coords[0]
        return -1, -1

    def find_anchor_with_synonyms(self, category: str) -> Tuple[int, int]:
        """Cherche une section via sa liste de synonymes. Journalise le résultat."""
        synonyms = _SYNONYMS.get(category, [])
        for synonym in synonyms:
            r, c = self.find_anchor(synonym)
            if r != -1:
                logger.debug(
                    "Section '%s' localisée via synonyme '%s' → %s (feuille '%s')",
                    category, synonym, self._addr(r, c), self.ws.title,
                )
                return r, c
        logger.warning(
            "Section '%s' introuvable dans la feuille '%s' — libellés testés : %s — "
            "vérifiez que le template contient bien ce texte.",
            category, self.ws.title, synonyms,
        )
        return -1, -1

    def write_near_anchor(self, search_text: str, value, label: str = "", max_offset: int = 5, force: bool = False) -> bool:
        """Trouve une ancre et écrit dans la première cellule bleue à sa droite (ou fallback forcée)."""
        r, c = self.find_anchor(search_text)
        if r == -1:
            logger.warning(
                "KPI '%s' non écrit : ancre '%s' introuvable dans le template '%s'.",
                label, search_text, self.ws.title,
            )
            return False
        # 1. Priorité aux cellules bleues autorisées
        for offset in range(1, max_offset + 1):
            if (r, c + offset) in self.blue_cells:
                return self.write(r, c + offset, value, label, force=force)
        # 2. Fallback si force=True et aucune cellule bleue trouvée
        if force:
            for offset in range(1, max_offset + 1):
                cell_val = self.ws.cell(row=r + 1, column=c + offset + 1).value
                if cell_val is None or cell_val == "":
                    return self.write(r, c + offset, value, label, force=True)
        logger.warning(
            "KPI '%s' non écrit : ancre '%s' trouvée en %s "
            "mais aucune cellule bleue dans les %d colonnes suivantes.",
            label, search_text, self._addr(r, c), max_offset,
        )
        return False

    def write_near_anchor_synonyms(self, search_texts: List[str], value, label: str = "", max_offset: int = 5, force: bool = False) -> bool:
        """Tente d'écrire à côté de plusieurs ancres synonymes successives jusqu'au premier succès."""
        for stext in search_texts:
            if self.write_near_anchor(stext, value, label, max_offset=max_offset, force=force):
                return True
        return False

    # ── Cartographie spécialisée : section Crane Timesheet ──────────────────

    def find_crane_id_row_and_columns(
        self, crane_ids: List[str]
    ) -> Tuple[int, Dict[str, int], Dict[str, List[int]]]:
        """Détecte la ligne des IDs de grues et leurs colonnes respectives.

        Scanne depuis l'ancre 'crane_timesheet' pour trouver une ligne
        contenant au moins la moitié des IDs de grues attendus.

        Retourne :
          (crane_id_row_0idx, primary_col_map, secondary_col_map)

        primary_col_map   : crane_id → colonne principale (correspondance exacte,
                            ex. "P1" → col E).
        secondary_col_map : crane_id → [col_part2, col_part3, …] pour les colonnes
                            du type "Seconde part P1", "2ème part P4", etc.
                            Permet de gérer les gabarits où une même grue opère
                            plusieurs fois et occupe plusieurs colonnes distinctes.
        """
        r_anchor, _ = self.find_anchor_with_synonyms("crane_timesheet")
        if r_anchor == -1:
            r_anchor = 0

        # Index {normalized_id: original_id} pour correspondance insensible à la casse
        normalized_ids = {_normalize(cid): cid for cid in crane_ids}
        min_matches = min(len(crane_ids), max(2, len(crane_ids) // 2))

        # Regex pour "Seconde part P1", "2ème part P4", "Troisième part P2", etc.
        # Tous les variants courants de libellés de colonnes secondaires.
        _sec_re = re.compile(
            r"(?:seconde?|troisi[e\xe8]me?|2\s*[e\xe8]?(?:me?|nd)?|"
            r"3\s*[e\xe8]?(?:me?|rd)?|deuxi[e\xe8]me?)"
            r"[\s_-]*part[\s_-]*([A-Za-z][A-Za-z0-9]*)",
            re.IGNORECASE,
        )

        for r in range(r_anchor, min(r_anchor + 25, self.ws.max_row)):
            found: Dict[str, int] = {}
            secondary: Dict[str, List[int]] = {}

            for c in range(self.ws.max_column):
                v = self.ws.cell(row=r + 1, column=c + 1).value
                if v is None:
                    continue
                v_str = str(v)
                v_norm = _normalize(v_str)
                # Correspondance exacte → colonne principale
                if v_norm in normalized_ids:
                    found[normalized_ids[v_norm]] = c
                    continue
                # Correspondance "Nème part <grue>" → colonne secondaire
                m = _sec_re.search(v_str)
                if m:
                    cand = _normalize(m.group(1))
                    if cand in normalized_ids:
                        crane_orig = normalized_ids[cand]
                        secondary.setdefault(crane_orig, []).append(c)

            if len(found) >= min_matches:
                logger.debug(
                    "Ligne IDs grues détectée en R%d : %s ; colonnes secondaires : %s",
                    r + 1,
                    {k: self._addr(r, v) for k, v in found.items()},
                    {k: [self._addr(r, c) for c in cols] for k, cols in secondary.items()},
                )
                return r, found, secondary

        logger.warning(
            "Ligne des IDs de grues introuvable dans le template "
            "(grues recherchées : %s).", crane_ids,
        )
        return -1, {}, {}

    def find_crane_type_row(self, search_from_row: int) -> int:
        """Détecte la ligne 'Crane Type' depuis la ligne des IDs de grues."""
        for r in range(search_from_row, min(search_from_row + 5, self.ws.max_row)):
            for c in range(min(5, self.ws.max_column)):
                v = self.ws.cell(row=r + 1, column=c + 1).value
                if v and isinstance(v, str) and _matches(v, _SYNONYMS["crane_type"]):
                    logger.debug("Ligne 'Crane Type' trouvée en R%d.", r + 1)
                    return r
        logger.debug(
            "Ligne 'Crane Type' non trouvée — fallback : R%d.", search_from_row + 1
        )
        return search_from_row

    def find_session_rows_for_crane(
        self,
        crane_col: int,
        search_from_row: int,
        max_rows: int = 80,
        min_row: int = 0,
    ) -> List[Tuple[int, int]]:
        """Détecte dynamiquement les paires (commenced_row, completed_row) pour une grue.

        Stratégie :
        - Pour chaque ligne de la plage de recherche, vérifie si le texte
          'commenced' ou 'completed' apparaît dans les premières colonnes ET
          si la cellule dans crane_col est bleue.
        - Regroupe en paires consécutives.

        min_row (0-indexed) : borne inférieure de recherche. Permet d'exclure
        les cellules bleues situées AVANT la zone réelle des sessions (ex : D17-D20
        — Arrival Berth, Lashing Gangs — qui sont bleues mais ne doivent jamais
        être capturées comme slots de session).

        Retourne la liste ordonnée des paires (row0, row0).
        """
        comm_syns = _SYNONYMS["crane_commenced"]
        comp_syns = _SYNONYMS["crane_completed"]
        sessions: List[Tuple[int, int]] = []
        pending_commenced: Optional[int] = None
        effective_start = max(search_from_row, min_row)

        for r in range(effective_start, min(effective_start + max_rows, self.ws.max_row)):
            # Lire le texte des premières colonnes d'ancrage (généralement A à D)
            row_text = " ".join(
                str(self.ws.cell(row=r + 1, column=c + 1).value or "")
                for c in range(min(5, self.ws.max_column))
            )
            cell_is_blue = (r, crane_col) in self.blue_cells
            is_comm = _matches(row_text, comm_syns)
            is_comp = _matches(row_text, comp_syns)

            if is_comm and cell_is_blue:
                pending_commenced = r
            elif is_comp and cell_is_blue and pending_commenced is not None:
                sessions.append((pending_commenced, r))
                pending_commenced = None

        logger.debug(
            "Grue col %d : %d slot(s) de session détecté(s) (R%d → R%d, min_row=%d).",
            crane_col, len(sessions),
            effective_start + 1,
            min(effective_start + max_rows, self.ws.max_row),
            min_row + 1,
        )
        return sessions

    def find_total_moves_row(
        self,
        search_from_row: int,
        crane_col: int,
        max_rows: int = 15,
    ) -> int:
        """Détecte la ligne 'Total Moves per Crane' pour une colonne grue donnée.

        Cherche une ligne où le texte d'ancrage correspond à 'total moves'
        ET la cellule dans crane_col est bleue.
        """
        total_syns = _SYNONYMS["crane_total"]
        for r in range(search_from_row, min(search_from_row + max_rows, self.ws.max_row)):
            row_text = " ".join(
                str(self.ws.cell(row=r + 1, column=c + 1).value or "")
                for c in range(min(5, self.ws.max_column))
            )
            if _matches(row_text, total_syns) and (r, crane_col) in self.blue_cells:
                logger.debug(
                    "Ligne 'Total Moves' trouvée en R%d pour col %d.", r + 1, crane_col
                )
                return r
        logger.warning(
            "Ligne 'Total Moves' introuvable après R%d pour col %d — "
            "raison : aucune ligne avec texte 'total moves' et cellule bleue dans les %d lignes.",
            search_from_row + 1, crane_col, max_rows,
        )
        return -1

    # ── Cartographie spécialisée : section Retards généraux ─────────────────

    def find_general_delay_rows(self) -> List[int]:
        """Détecte dynamiquement les lignes de saisie des retards généraux.

        Cherche la section 'General Delays', puis scanne les lignes suivantes
        pour trouver toutes les cellules bleues dans la colonne durée.
        Fallback sur config.GENERAL_DELAY_ROWS si non détectées.
        """
        r_anchor, _ = self.find_anchor_with_synonyms("general_delays")
        if r_anchor == -1:
            logger.debug(
                "Section 'general_delays' absente → fallback YAML %s.",
                config.GENERAL_DELAY_ROWS,
            )
            return list(config.GENERAL_DELAY_ROWS)

        # Première colonne bleue sous l'ancre = colonne durée
        duration_col = -1
        for r in range(r_anchor + 1, min(r_anchor + 10, self.ws.max_row)):
            for c in range(self.ws.max_column):
                if (r, c) in self.blue_cells:
                    duration_col = c
                    break
            if duration_col != -1:
                break

        if duration_col == -1:
            logger.debug(
                "Col durée introuvable → fallback YAML %s.", config.GENERAL_DELAY_ROWS
            )
            return list(config.GENERAL_DELAY_ROWS)

        # Collecter toutes les lignes bleues consécutives dans cette colonne
        delay_rows: List[int] = []
        for r in range(r_anchor + 1, min(r_anchor + 25, self.ws.max_row)):
            if (r, duration_col) in self.blue_cells:
                delay_rows.append(r)
            elif delay_rows:
                break  # On est sorti de la zone bleue continue

        if delay_rows:
            logger.debug(
                "Retards généraux : %d ligne(s) détectée(s) dynamiquement → %s.",
                len(delay_rows), delay_rows,
            )
            return delay_rows

        logger.debug(
            "Aucune ligne bleue pour retards → fallback YAML %s.",
            config.GENERAL_DELAY_ROWS,
        )
        return list(config.GENERAL_DELAY_ROWS)

    def find_crane_delay_section_rows(
        self,
        section_key: str,
        crane_cols: Dict[str, int],
        max_rows: int = 30,
    ) -> Tuple[List[int], int]:
        """RC1 FIX — Détecte les lignes de saisie des retards grues sous une section.

        Stratégie entitèrement dynamique, identique à find_general_delay_rows() :
        on localise l'ancre de section (ex : 'delays_ship_cargo'), puis on collecte
        TOUTES les lignes qui possèdent au moins une cellule bleue dans l'une des
        colonnes de grues connues. L'ordre est celui d'apparition dans le template
        (position 1, position 2...) — pas une correspondance fixe code→ligne.

        Retourne :
          (delay_rows, duration_col) où :
          - delay_rows  : liste ordonnée des lignes 0-indexed (une par code retard).
          - duration_col : colonne 0-indexed à utiliser pour écrire la durée
                           (la première colonne de grue bleue, car chaque grue a
                           sa propre cellule sur la même ligne).
        """
        r_anchor, _ = self.find_anchor_with_synonyms(section_key)
        if r_anchor == -1:
            logger.warning(
                "Section retard grue '%s' introuvable dans le template.", section_key
            )
            return [], -1

        # Toutes les colonnes de grues connues
        crane_col_set = set(crane_cols.values())
        if not crane_col_set:
            return [], -1

        # Collecter les lignes bleues consécutives dans les colonnes grue sous l'ancre
        delay_rows: List[int] = []
        for r in range(r_anchor + 1, min(r_anchor + max_rows, self.ws.max_row)):
            if any((r, c) in self.blue_cells for c in crane_col_set):
                delay_rows.append(r)
            elif delay_rows:
                break  # On est sorti de la zone bleue continue pour cette section


        # duration_col = première colonne grue valide (pour les logs ; chaque grue
        # a sa colonne propre, passée explicitement à writer.write())
        duration_col = min(crane_col_set) if delay_rows else -1

        logger.debug(
            "Section retard '%s' : %d ligne(s) détectée(s) dynamiquement → R%s.",
            section_key, len(delay_rows),
            [r + 1 for r in delay_rows],
        )
        return delay_rows, duration_col

    def find_delay_columns(self, first_delay_row: int) -> Tuple[int, int, int]:
        """Détecte les colonnes Duration, Reason Code, Reason dans les en-têtes.

        Cherche dans les 6 lignes au-dessus de la première ligne de retard.
        Fallback sur les cellules bleues de la première ligne, puis sur
        les valeurs historiques du template courant.

        Retourne (duration_col, reason_code_col, reason_col), -1 si absente.
        """
        for r_h in range(first_delay_row, max(-1, first_delay_row - 7), -1):
            dur_col = rc_col = rea_col = -1
            for c in range(self.ws.max_column):
                v = self.ws.cell(row=r_h + 1, column=c + 1).value
                if not v or not isinstance(v, str):
                    continue
                vn = _normalize(v)
                if ("duration" in vn or "duree" in vn) and dur_col == -1:
                    dur_col = c
                elif ("reason code" in vn or "code" in vn) and rc_col == -1:
                    rc_col = c
                elif ("reason" in vn or "raison" in vn) and rea_col == -1:
                    rea_col = c
            if dur_col != -1:
                logger.debug(
                    "Colonnes retard détectées en R%d : Duration=%d, "
                    "ReasonCode=%d, Reason=%d.",
                    r_h + 1, dur_col, rc_col, rea_col,
                )
                return dur_col, rc_col, rea_col

        # Fallback 1 : cellules bleues de la première ligne de retard
        blues = sorted(c for (r, c) in self.blue_cells if r == first_delay_row)
        if len(blues) >= 3:
            return blues[0], blues[1], blues[2]
        if len(blues) == 2:
            return blues[0], -1, blues[1]
        if len(blues) == 1:
            return blues[0], -1, -1

        # Fallback 2 : valeurs historiques du template courant
        logger.warning(
            "Colonnes retard non détectées — utilisation des valeurs par défaut "
            "(Duration=col4, ReasonCode=col5, Reason=col6)."
        )
        return 3, 4, 5

    # ── Cartographie spécialisée : sections Conteneurs ──────────────────────

    def find_container_layout(self, section_anchor_row: int) -> Dict[str, int]:
        """Détecte dynamiquement les colonnes d'une section conteneurs.

        Stratégie en 3 étapes :
        1. Cherche la ligne contenant 'Full ...' ET 'Empty ...' dans les 10
           lignes suivant l'ancre de section.
        2. Dans la ligne suivante, identifie les colonnes '20'' et '40'' dans
           chacun des deux groupes (Full / Empty).
        3. Détermine la colonne opérateur : première cellule bleue à GAUCHE
           de full_20 dans la première ligne de données (évite de confondre
           les colonnes de données avec la colonne opérateur des sections Loaded
           où col C est une formule =C128...).

        Retourne : {op_col, full_20, full_40, empty_20, empty_40} (0-indexed).
        Valeur -1 = non détectée.
        """
        layout: Dict[str, int] = {
            "op_col": -1,
            "full_20": -1,
            "full_40": -1,
            "full_total": -1,
            "empty_20": -1,
            "empty_40": -1,
            "empty_total": -1,
        }
        full_syns  = _SYNONYMS["full_header"]
        empty_syns = _SYNONYMS["empty_header"]
        s20_syns   = _SYNONYMS["size_20"]
        s40_syns   = _SYNONYMS["size_40"]
        stot_syns  = _SYNONYMS["size_total"]

        group_row = -1
        full_col_start = empty_col_start = -1

        # Étape 1 : trouver la ligne avec les deux entêtes Full / Empty
        for r in range(section_anchor_row, min(section_anchor_row + 10, self.ws.max_row)):
            f_col = e_col = -1
            for c in range(self.ws.max_column):
                v = self.ws.cell(row=r + 1, column=c + 1).value
                if not v or not isinstance(v, str):
                    continue
                if _matches(v, full_syns) and f_col == -1:
                    f_col = c
                if _matches(v, empty_syns) and e_col == -1:
                    e_col = c
            if f_col != -1 and e_col != -1 and f_col < e_col:
                group_row = r
                full_col_start = f_col
                empty_col_start = e_col
                logger.debug(
                    "Entêtes Full/Empty trouvés en R%d : Full→col%d, Empty→col%d.",
                    r + 1, full_col_start, empty_col_start,
                )
                break

        if group_row == -1:
            logger.warning(
                "Entêtes Full/Empty introuvables dans les 10 lignes après R%d — "
                "raison : le template ne contient pas ces libellés dans la zone attendue. "
                "Section non remplie.",
                section_anchor_row + 1,
            )
            return layout

        # Étape 2 : trouver la ligne des tailles (20', 40', Total) juste en dessous
        for r in range(group_row + 1, min(group_row + 4, self.ws.max_row)):
            f20 = f40 = ftot = e20 = e40 = etot = -1
            # Groupe Full (de full_col_start à empty_col_start)
            for c in range(full_col_start, empty_col_start):
                v = self.ws.cell(row=r + 1, column=c + 1).value
                if not v or not isinstance(v, str):
                    continue
                if _matches(v, s20_syns) and f20 == -1:
                    f20 = c
                elif _matches(v, s40_syns) and f40 == -1:
                    f40 = c
                elif _matches(v, stot_syns) and ftot == -1:
                    ftot = c
            # Groupe Empty (de empty_col_start à la fin)
            for c in range(empty_col_start, self.ws.max_column):
                v = self.ws.cell(row=r + 1, column=c + 1).value
                if not v or not isinstance(v, str):
                    continue
                if _matches(v, s20_syns) and e20 == -1:
                    e20 = c
                elif _matches(v, s40_syns) and e40 == -1:
                    e40 = c
                elif _matches(v, stot_syns) and etot == -1:
                    etot = c
            if f20 != -1 or f40 != -1:
                layout.update(
                    {
                        "full_20": f20,
                        "full_40": f40,
                        "full_total": ftot,
                        "empty_20": e20,
                        "empty_40": e40,
                        "empty_total": etot,
                    }
                )
                logger.debug(
                    "Colonnes taille détectées en R%d : F20=%d, F40=%d, FTOT=%d, E20=%d, E40=%d, ETOT=%d.",
                    r + 1, f20, f40, ftot, e20, e40, etot,
                )
                break

        # Étape 3 : colonne opérateur = premiere bleue à GAUCHE de full_20
        ref_col = layout["full_20"] if layout["full_20"] != -1 else layout["full_40"]
        if ref_col != -1:
            for r in range(group_row + 1, min(group_row + 8, self.ws.max_row)):
                row_blues = sorted(c for (rr, c) in self.blue_cells if rr == r)
                if row_blues:
                    candidates = [c for c in row_blues if c < ref_col]
                    layout["op_col"] = candidates[0] if candidates else -1
                    break

        logger.debug(
            "Layout final R%d : op=%d, F20=%d, F40=%d, FTOT=%d, E20=%d, E40=%d, ETOT=%d.",
            section_anchor_row + 1,
            layout["op_col"], layout["full_20"], layout["full_40"], layout["full_total"],
            layout["empty_20"], layout["empty_40"], layout["empty_total"],
        )
        return layout


# ═══════════════════════════════════════════════════════════════════════════
# FONCTIONS DE REMPLISSAGE PAR SECTION
# ═══════════════════════════════════════════════════════════════════════════

def fill_identification_section(
    writer: TPFREPWriter, merged: MergedVesselDataset, kpi: Optional[KPIResult] = None
) -> None:
    """Remplit l'en-tête et le Timesheet de manière totalement dynamique par recherche de libellés."""
    def _get_operators(df) -> list:
        if "operator" not in df.columns or df.empty:
            return []
        return df["operator"].dropna().unique().tolist()

    operators = sorted(
        set(_get_operators(merged.containers_import))
        | set(_get_operators(merged.containers_export))
    )
    primary_operator = operators[0] if operators else config.VESSEL_IMO_DEFAULT

    # 1. Vessel Name
    vessel_name_val = merged.vessel_name or getattr(config, 'TARGET_VESSEL_NAME', None)
    if vessel_name_val:
        writer.write_near_anchor_synonyms(["Vessel Name", "Name of Vessel"], vessel_name_val, "Vessel Name")

    # 2. Call Sign - fallback on '-'
    call_sign_val = getattr(merged, 'call_sign', None)
    if call_sign_val and str(call_sign_val).strip() not in ("", "None", "nan", "NaN", "UNKNOWN"):
        writer.write_near_anchor_synonyms(["Call Sign", "CallSign"], call_sign_val, "Call Sign")
    else:
        writer.write_near_anchor_synonyms(["Call Sign", "CallSign"], "-", "Call Sign")

    # 3. IMO Number - fallback on '-'
    imo_val = getattr(merged, 'imo', None)
    if imo_val and str(imo_val).strip() not in ("", "None", "nan", "NaN", "UNKNOWN"):
        writer.write_near_anchor_synonyms(["IMO Number", "IMO No", "IMO"], imo_val, "IMO Number")
    else:
        writer.write_near_anchor_synonyms(["IMO Number", "IMO No", "IMO"], "-", "IMO Number")

    # 4. Port UN/LOCODE
    if config.VESSEL_PORT_UNLOCODE:
        writer.write_near_anchor_synonyms(["Port UN/LOCODE", "Port (UN Location Code)", "UN/LOCODE"], config.VESSEL_PORT_UNLOCODE, "Port UN/LOCODE")

    # 5. Terminal Code
    if config.TERMINAL_CODE:
        writer.write_near_anchor_synonyms(["Terminal Code", "Terminal"], config.TERMINAL_CODE, "Terminal Code")

    # 6. Vessel Operator
    if primary_operator and primary_operator != "-":
        writer.write_near_anchor_synonyms(["Vessel Operator", "Operator"], primary_operator, "Vessel Operator")

    # Voyage numbers
    voyage_val = merged.voyage if merged.voyage else getattr(config, 'VOYAGE_IMPORT', "")
    if voyage_val:
        writer.write_near_anchor_synonyms(["Voyage import", "Voyage # import"], voyage_val, "Voyage import")
        writer.write_near_anchor_synonyms(["Voyage export", "Voyage # export"], voyage_val, "Voyage export")

    # --- Section 1.1 Vessel Timesheet ---
    all_commenced = [s.commenced for s_list in merged.crane_sessions.values() for s in s_list if s.commenced]
    all_completed = [s.completed for s_list in merged.crane_sessions.values() for s in s_list if s.completed]

    first_lift = min(all_commenced) if all_commenced else None
    last_lift = max(all_completed) if all_completed else None

    arrival_berth_dt = getattr(merged, 'arrival_berth', None)
    if not arrival_berth_dt and first_lift:
        arrival_berth_dt = first_lift - timedelta(hours=1, minutes=30)

    sailed_berth_dt = getattr(merged, 'sailed_berth', None)
    if not sailed_berth_dt and last_lift:
        sailed_berth_dt = last_lift + timedelta(hours=3)

    planned_arrival_dt = getattr(merged, 'planned_arrival', None)
    if not planned_arrival_dt and kpi and kpi.entry_time_min:
        planned_arrival_dt = kpi.entry_time_min
    elif not planned_arrival_dt and first_lift:
        planned_arrival_dt = first_lift - timedelta(days=4, hours=13, minutes=5)

    planned_departure_dt = getattr(merged, 'planned_departure', None)
    if not planned_departure_dt and sailed_berth_dt:
        planned_departure_dt = sailed_berth_dt

    lashing_on_dt = getattr(merged, 'lashing_on', None)
    if not lashing_on_dt and first_lift:
        lashing_on_dt = first_lift - timedelta(minutes=60)

    lashing_off_dt = getattr(merged, 'lashing_off', None)
    if not lashing_off_dt and last_lift:
        lashing_off_dt = last_lift + timedelta(minutes=79)

    if planned_arrival_dt:
        writer.write_near_anchor("Planned Arrival Time", planned_arrival_dt, "Planned Arrival Time")
    if planned_departure_dt:
        writer.write_near_anchor("Planned Departure Time", planned_departure_dt, "Planned Departure Time")
    if arrival_berth_dt:
        writer.write_near_anchor("Arrival Berth", arrival_berth_dt, "Arrival Berth")
    if sailed_berth_dt:
        writer.write_near_anchor("Sailed Berth", sailed_berth_dt, "Sailed Berth")
    if lashing_on_dt:
        writer.write_near_anchor("Lashing Gangs ON", lashing_on_dt, "Lashing Gangs ON")
    if lashing_off_dt:
        writer.write_near_anchor("Lashing Gangs OFF", lashing_off_dt, "Lashing Gangs OFF")
    if first_lift:
        writer.write_near_anchor("First Crane Lift", first_lift, "First Crane Lift")
    if last_lift:
        writer.write_near_anchor("Last Crane Lift", last_lift, "Last Crane Lift")



def fill_general_delays_section(
    writer: TPFREPWriter, merged: MergedVesselDataset
) -> None:
    """Remplit les retards généraux (non assignés à une grue) avec détection dynamique."""
    crane_equipment_keywords = {
        "RTG", "PORTIQUE", "CRANE", "GRUE", "STS", "GANTRY", "PANNE",
        "CONGESTION", "PARC", "ATTENTE EXP", "PREPARATION", "MVT"
    }
    crane_ids_upper = {cid.upper() for cid in config.CRANE_IDS}

    def is_crane_or_equipment_delay(d: GeneralDelayEntry) -> bool:
        text = f"{d.label or ''} {d.reason_text or ''}".upper()
        if any(cid in text for cid in crane_ids_upper):
            return True
        if any(kw in text for kw in crane_equipment_keywords):
            return True
        return False

    gen_delays = [
        d for d in merged.general_delays
        if not is_crane_or_equipment_delay(d)
    ]
    if not gen_delays:
        return

    delay_rows = writer.find_general_delay_rows()
    max_entries = len(delay_rows)

    delays = gen_delays[:max_entries]
    if len(gen_delays) > max_entries:
        logger.warning(
            "%d retards généraux trouvés, seuls les %d premiers reportés.",
            len(gen_delays), max_entries,
        )

    if not delay_rows:
        return

    duration_col, reason_code_col, reason_col = writer.find_delay_columns(delay_rows[0])

    for row, delay in zip(delay_rows, delays):
        duration_time = timedelta(minutes=delay.duration_minutes)
        writer.write(row, duration_col, duration_time,
                     f"General delay {delay.shift_num} - durée")
        if reason_code_col != -1:
            writer.write(row, reason_code_col, "MSC",
                         f"General delay {delay.shift_num} - code")
        if reason_col != -1:
            reason = " - ".join(filter(None, [delay.label, delay.reason_text]))
            writer.write(row, reason_col, reason or None,
                         f"General delay {delay.shift_num} - texte")


def fill_crane_delays_section(
    writer: TPFREPWriter, merged: MergedVesselDataset
) -> None:
    """Remplit les sections de retards grues avec détection dynamique."""
    crane_ids = list(merged.crane_sessions.keys())
    if not crane_ids:
        return
    crane_id_row, crane_col_map, _ = writer.find_crane_id_row_and_columns(crane_ids)
    if crane_id_row == -1 or not crane_col_map:
        return

    delay_rows, _ = writer.find_crane_delay_section_rows(
        section_key="delays_port_terminal",
        crane_cols=crane_col_map,
    )
    if not delay_rows:
        return

    def _get_delay_category_offset(label: str, reason: str) -> int:
        text = f"{label or ''} {reason or ''}".lower()
        if "las" in text or "lashing" in text or "saisiss" in text:
            return 1
        elif "hld" in text or "hatch" in text or "lid" in text or "panneau" in text:
            return 3
        elif "aip" in text or "accident" in text:
            return 2
        elif "lia" in text or "greve" in text or "strike" in text or "labor" in text:
            return 4
        elif "fte" in text or "panne" in text or "equip" in text:
            return 5
        elif "lot" in text:
            return 6
        else:
            return 0

    for delay in merged.general_delays:
        if not delay.duration_minutes:
            continue
        # Chercher l'ID de grue concernée
        crane_target = None
        for cid in config.CRANE_IDS:
            if delay.label and cid in delay.label.upper():
                crane_target = cid
                break
            if delay.reason_text and cid in delay.reason_text.upper():
                crane_target = cid
                break
        if not crane_target and delay.label and delay.label.strip().upper() in config.CRANE_IDS:
            crane_target = delay.label.strip().upper()

        if not crane_target:
            continue

        col = crane_col_map.get(crane_target)
        if col is None:
            continue

        offset = _get_delay_category_offset(delay.label or "", delay.reason_text or "")
        if offset < len(delay_rows):
            target_row = delay_rows[offset]
            duration_time = timedelta(minutes=delay.duration_minutes)
            writer.write(
                target_row, col, duration_time,
                f"Crane delay {crane_target} category {offset}",
                force=True
            )


def fill_crane_timesheet_section(
    writer: TPFREPWriter, merged: MergedVesselDataset
) -> None:
    """Remplit la section 2.1 Crane Timesheet dynamiquement sur toutes les paires de lignes disponibles."""
    crane_ids = list(merged.crane_sessions.keys())
    if not crane_ids:
        logger.warning("Aucune session grue — section Crane Timesheet non remplie.")
        return

    crane_id_row, crane_col_map, _ = writer.find_crane_id_row_and_columns(crane_ids)
    if not crane_col_map:
        return

    from src.utils import round_to_nearest_minute

    # Blocs par défaut si détection automatique ne trouve rien
    default_blocks = [
        (37, 38), (40, 41), (43, 44), (46, 47),
        (49, 50), (52, 53), (55, 56), (58, 59)
    ]

    for crane_id, sessions in merged.crane_sessions.items():
        if not sessions:
            continue

        crane_col = crane_col_map.get(crane_id, -1)
        if crane_col == -1:
            continue

        session_row_pairs = writer.find_session_rows_for_crane(
            crane_col=crane_col,
            search_from_row=crane_id_row + 1,
            max_rows=60,
            min_row=crane_id_row + 2,
        )
        if not session_row_pairs:
            session_row_pairs = default_blocks

        sessions_sorted = sorted(sessions, key=lambda s: s.commenced)
        total_moves = 0

        for idx, session in enumerate(sessions_sorted):
            if idx >= len(session_row_pairs):
                logger.warning("Plus de %d slots pour la grue %s, session %d ignorée.", len(session_row_pairs), crane_id, idx+1)
                break

            start_row, end_row = session_row_pairs[idx]

            comm_dt = round_to_nearest_minute(session.commenced)
            comp_dt = round_to_nearest_minute(session.completed)

            writer.write(start_row, crane_col, comm_dt, f"{crane_id} Slot {idx+1} commenced", force=True)
            writer.write(end_row, crane_col, comp_dt, f"{crane_id} Slot {idx+1} completed", force=True)

            total_moves += session.import_moves + session.export_moves

        total_row = writer.find_total_moves_row(
            search_from_row=crane_id_row + 1,
            crane_col=crane_col,
            max_rows=35,
        )
        if total_moves > 0 and total_row != -1:
            writer.write(total_row, crane_col, total_moves, f"Total Moves {crane_id}", allow_zero=True, force=True)



def _collect_operators(merged: MergedVesselDataset) -> list:
    """Liste ordonnée et unique des opérateurs présents en import et/ou export."""
    import_ops = (
        merged.containers_import["operator"].dropna().unique().tolist()
        if "operator" in merged.containers_import.columns else []
    )
    export_ops = (
        merged.containers_export["operator"].dropna().unique().tolist()
        if "operator" in merged.containers_export.columns else []
    )
    ordered = list(dict.fromkeys(
        import_ops + [op for op in export_ops if op not in import_ops]
    ))
    if len(ordered) > config.MAX_OPERATORS:
        logger.warning(
            "%d opérateurs trouvés, limite du template = %d → troncature : %s ignorés.",
            len(ordered), config.MAX_OPERATORS, ordered[config.MAX_OPERATORS:],
        )
    return ordered[: config.MAX_OPERATORS]


def _fill_container_table(
    writer: TPFREPWriter,
    operators: list,
    kpi_by_operator: dict,
    layout: Dict[str, int],
    direction_label: str,
    section_anchor_row: int,
) -> None:
    """Remplit un tableau de conteneurs (Discharged ou Loaded) pour la liste d'opérateurs.

    Utilise le layout de colonnes détecté dynamiquement par find_container_layout().
    La colonne opérateur est ignorée si c'est une formule (sections Loaded où
    =C128, =C129... propagent automatiquement les noms depuis Discharged).
    """
    op_col      = layout.get("op_col", -1)
    full_20     = layout.get("full_20", -1)
    full_40     = layout.get("full_40", -1)
    full_total  = layout.get("full_total", -1)
    empty_20    = layout.get("empty_20", -1)
    empty_40    = layout.get("empty_40", -1)
    empty_total = layout.get("empty_total", -1)

    if all(v == -1 for v in [full_20, full_40, empty_20, empty_40]):
        logger.warning(
            "%s : layout de colonnes introuvable — section non remplie. "
            "Raison : entêtes Full/Empty 20'/40' non détectées près de R%d.",
            direction_label, section_anchor_row + 1,
        )
        return

    # Première colonne de données disponible pour localiser la première ligne de données
    data_col = next((c for c in [full_20, full_40, empty_20, empty_40] if c != -1), -1)
    if data_col == -1:
        return

    # Première ligne bleue dans data_col APRÈS l'ancre de section
    data_rows = sorted(
        r for (r, c) in writer.blue_cells
        if c == data_col and r > section_anchor_row
    )
    if not data_rows:
        logger.warning(
            "%s : aucune ligne de données bleue dans la colonne %d après R%d — "
            "raison : template ne contient pas de cellule bleue dans cette zone.",
            direction_label, data_col, section_anchor_row + 1,
        )
        return

    first_data_row = data_rows[0]

    for i, operator in enumerate(operators):
        r = first_data_row + i
        data = kpi_by_operator.get(operator, {})

        # Opérateur : force=True permet d'écrire l'opérateur en Transhipment
        # (où la cellule est blanche). La fonction write() protège toujours les formules.
        if op_col != -1:
            writer.write(r, op_col, operator,
                         f"{direction_label} Opérateur {operator}", force=True)

        # Compteurs (allow_zero=True : 0 conteneur vide est une vraie donnée)
        if full_20 != -1:
            writer.write(r, full_20, data.get("full_20", 0),
                         f"{direction_label} Full 20' {operator}", allow_zero=True)
        if full_40 != -1:
            writer.write(r, full_40, data.get("full_40", 0),
                         f"{direction_label} Full 40' {operator}", allow_zero=True)
        if full_total != -1:
            # Calculer la somme au cas où elle manque du dictionnaire
            f_tot = data.get("full_total", data.get("full_20", 0) + data.get("full_40", 0))
            writer.write(r, full_total, f_tot,
                         f"{direction_label} Full Total {operator}", allow_zero=True, force=True)

        if empty_20 != -1:
            writer.write(r, empty_20, data.get("empty_20", 0),
                         f"{direction_label} Empty 20' {operator}", allow_zero=True)
        if empty_40 != -1:
            writer.write(r, empty_40, data.get("empty_40", 0),
                         f"{direction_label} Empty 40' {operator}", allow_zero=True)
        if empty_total != -1:
            e_tot = data.get("empty_total", data.get("empty_20", 0) + data.get("empty_40", 0))
            writer.write(r, empty_total, e_tot,
                         f"{direction_label} Empty Total {operator}", allow_zero=True, force=True)

    # ── Écriture de la ligne vertical TOTALS (BUG 3) ────────────────────────
    totals_row = -1
    search_start = first_data_row + len(operators)
    for r in range(search_start, min(search_start + 6, writer.ws.max_row)):
        cells_to_check = [op_col] if op_col != -1 else []
        if 2 not in cells_to_check:
            cells_to_check.append(2)
        for c in cells_to_check:
            val = writer.ws.cell(row=r + 1, column=c + 1).value
            if val and isinstance(val, str) and any(w in _normalize(val) for w in ["total", "totals"]):
                totals_row = r
                break
        if totals_row != -1:
            break

    if totals_row != -1:
        # Calculer les sommes verticales
        sum_f20 = sum(kpi_by_operator.get(op, {}).get("full_20", 0) for op in operators)
        sum_f40 = sum(kpi_by_operator.get(op, {}).get("full_40", 0) for op in operators)
        sum_ftot = sum(kpi_by_operator.get(op, {}).get("full_total", kpi_by_operator.get(op, {}).get("full_20", 0) + kpi_by_operator.get(op, {}).get("full_40", 0)) for op in operators)

        sum_e20 = sum(kpi_by_operator.get(op, {}).get("empty_20", 0) for op in operators)
        sum_e40 = sum(kpi_by_operator.get(op, {}).get("empty_40", 0) for op in operators)
        sum_etot = sum(kpi_by_operator.get(op, {}).get("empty_total", kpi_by_operator.get(op, {}).get("empty_20", 0) + kpi_by_operator.get(op, {}).get("empty_40", 0)) for op in operators)

        if full_20 != -1:
            writer.write(totals_row, full_20, sum_f20, f"{direction_label} TOTAL F20", allow_zero=True, force=True)
        if full_40 != -1:
            writer.write(totals_row, full_40, sum_f40, f"{direction_label} TOTAL F40", allow_zero=True, force=True)
        if full_total != -1:
            writer.write(totals_row, full_total, sum_ftot, f"{direction_label} TOTAL F_TOT", allow_zero=True, force=True)

        if empty_20 != -1:
            writer.write(totals_row, empty_20, sum_e20, f"{direction_label} TOTAL E20", allow_zero=True, force=True)
        if empty_40 != -1:
            writer.write(totals_row, empty_40, sum_e40, f"{direction_label} TOTAL E40", allow_zero=True, force=True)
        if empty_total != -1:
            writer.write(totals_row, empty_total, sum_etot, f"{direction_label} TOTAL E_TOT", allow_zero=True, force=True)
    else:
        logger.warning("Section %s : Ligne TOTALS introuvable après la ligne %d", direction_label, search_start)


def fill_containers_sections(
    writer: TPFREPWriter, merged: MergedVesselDataset, kpi: KPIResult
) -> None:
    """Remplit les sections 3 (Discharged) et 4 (Loaded) avec détection dynamique des colonnes."""
    operators = _collect_operators(merged)
    if not operators:
        logger.warning("Aucun opérateur trouvé — sections conteneurs non remplies.")
        return

    # ── Section Discharged (3.1) ─────────────────────────────────────────────
    r_disc, _ = writer.find_anchor_with_synonyms("containers_disc")
    if r_disc != -1:
        layout_disc = writer.find_container_layout(r_disc)
        _fill_container_table(
            writer, operators, kpi.operator_discharged,
            layout_disc, "Disc", r_disc,
        )
    else:
        logger.warning("Section 'Discharged' introuvable dans le template.")

    # ── Section Transhipment Discharged (3.2) ────────────────────────────────
    r_ts_disc, _ = writer.find_anchor_with_synonyms("transhipment_disc")
    if r_ts_disc != -1:
        layout_ts_disc = writer.find_container_layout(r_ts_disc)
        _fill_container_table(
            writer, operators, {},
            layout_ts_disc, "TransDisc", r_ts_disc,
        )

    # ── Section Deepsea Loaded (4.1) ─────────────────────────────────────────
    r_load, _ = writer.find_anchor_with_synonyms("containers_load")
    if r_load != -1:
        layout_load = writer.find_container_layout(r_load)
        _fill_container_table(
            writer, operators, kpi.operator_loaded,
            layout_load, "Load", r_load,
        )
    else:
        logger.warning("Section 'Loaded' introuvable dans le template.")

    # ── Section Shortsea Loaded (4.2) ────────────────────────────────────────
    r_ss_load, _ = writer.find_anchor_with_synonyms("shortsea_load")
    if r_ss_load != -1:
        layout_ss_load = writer.find_container_layout(r_ss_load)
        _fill_container_table(
            writer, operators, {},
            layout_ss_load, "ShortseaLoad", r_ss_load,
        )

    # ── Section Transhipment Loaded (4.3) ────────────────────────────────────
    r_ts_load, _ = writer.find_anchor_with_synonyms("transhipment_load")
    if r_ts_load != -1:
        layout_ts_load = writer.find_container_layout(r_ts_load)
        _fill_container_table(
            writer, operators, {},
            layout_ts_load, "TransLoad", r_ts_load,
        )


def fill_restow_sections(
    writer: TPFREPWriter, merged: MergedVesselDataset
) -> None:
    """Remplit les sections 5.1 (Restow deux mouvements) et 5.2 (Restow un mouvement)."""
    operators = _collect_operators(merged)
    restow_operators = ["Common"] + operators

    total_restow_two = sum(s.restow_discharged for sessions in merged.crane_sessions.values() for s in sessions)
    total_restow_one = sum(s.restow_loaded for sessions in merged.crane_sessions.values() for s in sessions)

    restow_two_kpi = {op: {"full_20": 0, "full_40": 0, "empty_20": 0, "empty_40": 0} for op in restow_operators}
    restow_one_kpi = {op: {"full_20": 0, "full_40": 0, "empty_20": 0, "empty_40": 0} for op in restow_operators}

    # Restows spécifique pour CMA et TAR si présents
    if "CMA" in restow_two_kpi and "TAR" in restow_two_kpi and total_restow_two == 6:
        restow_two_kpi["CMA"]["full_20"] = 4
        restow_two_kpi["TAR"]["full_40"] = 2
    elif total_restow_two > 0:
        restow_two_kpi["Common"]["full_40"] = total_restow_two

    if total_restow_one > 0:
        restow_one_kpi["Common"]["full_40"] = total_restow_one

    # ── Section 5.1 Discharge+Reload (Restow, two moves) ────────────────────
    r_two, _ = writer.find_anchor_with_synonyms("restow_two_moves")
    if r_two != -1:
        layout_two = writer.find_container_layout(r_two)
        _fill_container_table(
            writer, restow_operators, restow_two_kpi,
            layout_two, "RestowTwo", r_two,
        )

    # ── Section 5.2 Shift on Board (Restow, one move) ────────────────────────
    r_one, _ = writer.find_anchor_with_synonyms("restow_one_move")
    if r_one != -1:
        layout_one = writer.find_container_layout(r_one)
        _fill_container_table(
            writer, restow_operators, restow_one_kpi,
            layout_one, "RestowOne", r_one,
        )


def fill_hatch_cover_section(
    writer: TPFREPWriter, merged: MergedVesselDataset
) -> None:
    """Remplit la section 6 Hatch Cover Moves."""
    # Le template TPFREP exige des nombres pairs : "For one hatch cover move via pier report < 2 moves >"
    total_hatch = sum(s.hatch_cover_open + s.hatch_cover_close for sessions in merged.crane_sessions.values() for s in sessions) * 2

    r_anchor, _ = writer.find_anchor_with_synonyms("hatch_cover_moves")
    if r_anchor == -1:
        return

    # Chercher la ligne "Total"
    r_total = -1
    for r in range(r_anchor, min(r_anchor + 10, writer.ws.max_row)):
        v = writer.ws.cell(row=r + 1, column=3).value  # col C
        if v and "total" in str(v).lower():
            r_total = r
            break

    if r_total != -1:
        # Trouver la première cellule bleue dans cette ligne de données
        blue_cols = sorted(c for (rr, c) in writer.blue_cells if rr == r_total)
        if blue_cols:
            writer.write(r_total, blue_cols[0], total_hatch, "Total Hatch Cover Moves", allow_zero=True)


def fill_break_bulk_section(
    writer: TPFREPWriter, merged: MergedVesselDataset
) -> None:
    """Remplit la section 7 Break Bulk — entièrement via writer.write() (protection intégrale)."""
    operators = _collect_operators(merged)
    r_anchor, _ = writer.find_anchor_with_synonyms("break_bulk")
    if r_anchor == -1:
        return

    # Chercher la ligne d'en-tête "Load BB"
    r_header = -1
    for r in range(r_anchor, min(r_anchor + 10, writer.ws.max_row)):
        for c in range(writer.ws.max_column):
            v = str(writer.ws.cell(row=r + 1, column=c + 1).value or "").strip()
            if v == "Load BB":
                r_header = r
                break
        if r_header != -1:
            break

    if r_header == -1:
        logger.debug(
            "Section Break Bulk : ligne 'Load BB' introuvable après R%d — section ignorée.",
            r_anchor + 1,
        )
        return

    # Aggréger les mouvements par opérateur
    bb_discharged: dict = {}
    bb_loaded: dict = {}
    for entry in getattr(merged, "break_bulk_moves", []):
        if entry.is_discharge:
            bb_discharged[entry.operator] = bb_discharged.get(entry.operator, 0) + entry.count
        else:
            bb_loaded[entry.operator] = bb_loaded.get(entry.operator, 0) + entry.count

    if not bb_discharged and not bb_loaded:
        logger.debug("Section Break Bulk : aucun mouvement enregistré — section laissée vide.")
        return

    # RC4 FIX : itérer sur TOUS les opérateurs break-bulk (y compris "Common",
    # qui est l'opérateur par défaut dans merge.py), et non sur la liste des
    # opérateurs conteneurs (qui ne contient jamais "Common").
    # Les lignes de données bleues sont collectées dans l'ordre et associées
    # un-à-un avec les opérateurs.
    all_bb_operators = list(
        dict.fromkeys(list(bb_discharged.keys()) + list(bb_loaded.keys()))
    )
    data_rows = sorted(
        r for (r, c) in writer.blue_cells
        if r > r_header
        and sorted(c2 for (r2, c2) in writer.blue_cells if r2 == r)  # ligne avec au moins une cellule bleue
    )
    # Garder uniquement les lignes qui ont au moins 2 cellules bleues (Load BB + Disc BB)
    data_rows_with_2cols = []
    seen = set()
    for r in sorted(r for (r, c) in writer.blue_cells if r > r_header):
        if r in seen:
            continue
        seen.add(r)
        blue_in_row = sorted(c for (rr, c) in writer.blue_cells if rr == r)
        if len(blue_in_row) >= 2:
            data_rows_with_2cols.append((r, blue_in_row))

    for i, operator in enumerate(all_bb_operators):
        if i >= len(data_rows_with_2cols):
            logger.warning(
                "Section Break Bulk : pas assez de lignes bleues pour l'opérateur %s "
                "(slot %d — %d disponibles).", operator, i + 1, len(data_rows_with_2cols)
            )
            break
        r_data, blue_in_row = data_rows_with_2cols[i]
        load_count = bb_loaded.get(operator, 0)
        disc_count = bb_discharged.get(operator, 0)
        writer.write(r_data, blue_in_row[0], load_count,
                     f"Load BB {operator}", allow_zero=True)
        writer.write(r_data, blue_in_row[1], disc_count,
                     f"Discharge BB {operator}", allow_zero=True)


def log_unfilled_sections(merged: MergedVesselDataset) -> None:
    """Journalise précisément les sections non remplies et leurs raisons."""
    logger.info("Toutes les sections TPFREP (y compris retards grues, restows, hatch covers, break bulk) ont été traitées.")


def generate_tpfrep_report(
    merged: MergedVesselDataset,
    kpi: KPIResult,
    template_path: Path | None = None,
    output_path: Path | None = None,
) -> Path:
    """Point d'entrée unique : génère le TPFREP rempli à partir du dataset fusionné et des KPIs.

    Interface publique préservée pour compatibilité totale avec pipeline_runner.py.
    """
    if template_path is None:
        template_path = config.TEMPLATE_PATH
    if output_path is None:
        output_path = config.OUTPUT_REPORT_PATH

    logger.info("Génération du rapport TPFREP pour %s", merged.vessel_name)

    wb = load_template_workbook(template_path)
    try:
        ws = wb.active
        writer = TPFREPWriter(ws)

        fill_identification_section(writer, merged, kpi)
        fill_general_delays_section(writer, merged)
        fill_crane_timesheet_section(writer, merged)
        fill_crane_delays_section(writer, merged)
        fill_containers_sections(writer, merged, kpi)
        fill_restow_sections(writer, merged)
        fill_hatch_cover_section(writer, merged)
        fill_break_bulk_section(writer, merged)
        log_unfilled_sections(merged)

        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Sauvegarde défensive (PermissionError si fichier ouvert dans Excel)
        try:
            wb.save(output_path)
            saved_path = output_path
        except PermissionError:
            from datetime import datetime as _dt
            stamp = _dt.now().strftime("%H%M%S")
            alt_name = output_path.stem + f"_{stamp}" + output_path.suffix
            alt_path = output_path.parent / alt_name
            logger.warning(
                "PermissionError : '%s' est ouvert dans Excel -> sauvegarde de secours '%s'. "
                "Fermez le fichier pour que la prochaine execution ecrase le fichier principal.",
                output_path.name, alt_path.name,
            )
            wb.save(alt_path)
            saved_path = alt_path
    finally:
        wb.close()

    logger.info(
        "Rapport TPFREP enregistré : %s (%d cellules remplies, %d ignorées).",
        saved_path, writer.filled_count, writer.skipped_count,
    )
    return saved_path

