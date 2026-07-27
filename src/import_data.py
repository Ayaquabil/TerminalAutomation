"""
src/import_data.py — Chargement automatique de tous les classeurs Excel
depuis data/input/ (rapports de shift + IMPORT/EXPORT MASTERYD) et
data/template/ (template TPFREP).

DÉCOUVERTE PAR CONTENU (et non par nom de fichier).

CORRECTIFS v2 (programmation défensive) :
  - BUG 3 : _earliest_doc_time() détecte la colonne DOC dynamiquement.
  - BUG 5 : SHIFT_NUMBER_RE assouplie pour "1er shift", "2ème shift", etc.
"""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import openpyxl

import config
from src.logger import get_logger
from src.settings_loader import get_settings

logger = get_logger("import_data")


@dataclass
class RawShiftReport:
    shift_num: int
    file_path: Path
    sheet_title: str
    rows: list = field(default_factory=list)


@dataclass
class RawMasterydFile:
    direction: str
    file_path: Path
    header: tuple
    rows: list = field(default_factory=list)


class InputDiscoveryError(RuntimeError):
    """Levée quand un fichier d'entrée attendu n'est pas trouvé ou pas identifiable."""


@dataclass
class EscaleInfo:
    name: str
    min_date: Optional[datetime]
    max_date: Optional[datetime]
    container_count: int


class MultipleEscalesError(RuntimeError):
    """Levée quand plusieurs escales sont détectées et qu'aucune n'est sélectionnée."""
    def __init__(self, message: str, escales: list[EscaleInfo]):
        super().__init__(message)
        self.escales = escales


BLUE_CELL_TEMPLATE_THRESHOLD = get_settings().get(
    "file_detection", "blue_cell_template_threshold", default=100
)


def _cell_text_contains(rows, needle: str, max_rows: int = 20) -> bool:
    needle_lower = needle.lower()
    for row in rows[:max_rows]:
        for cell in row:
            if isinstance(cell, str) and needle_lower in cell.strip().lower():
                return True
    return False


def _count_blue_cells(ws, limit: int = BLUE_CELL_TEMPLATE_THRESHOLD + 1) -> int:
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            fill = cell.fill
            if fill is not None and getattr(fill, "patternType", None) == "solid":
                fg = fill.fgColor
                if fg is not None:
                    is_blue = getattr(fg, "indexed", None) == config.BLUE_INDEXED or (
                        getattr(fg, "type", None) == "rgb"
                        and str(getattr(fg, "rgb", "")).upper() == config.BLUE_RGB_FALLBACK
                    )
                    if is_blue:
                        count += 1
                        if count >= limit:
                            return count
    return count


def looks_like_template(path: Path) -> bool:
    try:
        wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    except Exception as exc:
        logger.debug("Impossible d'ouvrir %s pour sniff template : %s", path.name, exc)
        return False
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True, max_row=10))
    has_title_text = _cell_text_contains(rows, "TERMINAL DEPARTURE AND PERFORMANCE REPORT")
    blue_count = _count_blue_cells(ws) if not has_title_text else 0
    wb.close()
    return has_title_text or blue_count >= BLUE_CELL_TEMPLATE_THRESHOLD


def looks_like_shift_report(rows) -> bool:
    for row in rows[:15]:
        for cell in row:
            if isinstance(cell, str):
                c_str = cell.strip().lower()
                if "portique" in c_str or "grue" in c_str or "crane" in c_str:
                    return True
    return False


def looks_like_masteryd(header) -> bool:
    header_clean = [str(h).strip().upper() if h is not None else "" for h in header]
    has_container = any("CONTENEUR" in h or "CONTAINER" in h for h in header_clean)
    has_operator = any("EXPLOITANT" in h or "OPERATOR" in h or "ARMATEUR" in h for h in header_clean)
    return has_container and has_operator


def detect_masteryd_direction(header: tuple, rows: list) -> Optional[str]:
    header_clean = [str(h).strip().upper() if h is not None else "" for h in header]
    idx = -1
    for i, h in enumerate(header_clean):
        if "EXP IMP" in h or "DIRECTION" in h or "FLUX" in h or "TRB" in h:
            idx = i
            break
    if idx == -1:
        return None
        
    values = [str(r[idx]).strip().upper() for r in rows if idx < len(r) and r[idx]]
    if not values:
        return None
    i_count = values.count("I")
    e_count = values.count("E")
    if i_count == 0 and e_count == 0:
        return None
    return "IMPORT" if i_count >= e_count else "EXPORT"


# BUG 5 CORRIGÉ — SHIFT_NUMBER_RE assouplie
SHIFT_NUMBER_RE = re.compile(
    r"shift\D{0,5}(\d+)"
    r"|(\d+)\D{0,5}(?:i?[eè]me?|er?)?"
    r"\s*(?:shift|rapport)",
    re.IGNORECASE,
)


def detect_shift_number(
    rows,
    path: Optional[Path] = None,
    sheet_title: Optional[str] = None,
    max_rows: int = 15,
) -> Optional[int]:
    logger.debug("Détection numéro shift (nom fichier, onglet, puis regex cellules)")
    
    # 1. Essayer le nom de l'onglet/feuille
    if sheet_title:
        m = SHIFT_NUMBER_RE.search(sheet_title)
        if m:
            digit = next((g for g in m.groups() if g is not None), None)
            if digit:
                return int(digit)

    # 2. Essayer le nom du fichier
    if path and path.name:
        m = SHIFT_NUMBER_RE.search(path.name)
        if m:
            digit = next((g for g in m.groups() if g is not None), None)
            if digit:
                return int(digit)

    # 3. Essayer sur les premières lignes
    for row in rows[:max_rows]:
        for cell in row:
            if not isinstance(cell, str):
                continue
            m = SHIFT_NUMBER_RE.search(cell)
            if m:
                digit = next((g for g in m.groups() if g is not None), None)
                if digit:
                    return int(digit)
    return None


def _find_doc_col_index(header_row) -> Optional[int]:
    for idx, cell in enumerate(header_row):
        if cell is not None and "doc" in str(cell).strip().lower():
            return idx
    return None


# BUG 3 CORRIGÉ — _earliest_doc_time() dynamique
def _earliest_doc_time(rows) -> Optional[tuple]:
    from datetime import time as dtime
    logger.debug("Tri chronologique shifts — recherche colonne DOC dynamique")
    header_idx = None
    for i, row in enumerate(rows[:20]):
        if any(isinstance(c, str) and "portiques" in c.strip().lower()
               for c in row if c is not None):
            header_idx = i
            break
    if header_idx is None:
        return None
    subheader_row = rows[header_idx + 1] if header_idx + 1 < len(rows) else None
    doc_col = None
    if subheader_row:
        doc_col = _find_doc_col_index(subheader_row)
    if doc_col is None:
        doc_col = config.SHIFT_COL_DOC
        logger.debug("_earliest_doc_time : repli sur index fixe %d.", doc_col)
    for row in rows[header_idx + 2: header_idx + 2 + 6]:
        if len(row) > doc_col:
            val = row[doc_col]
            if isinstance(val, dtime) and not (val.hour == 0 and val.minute == 0):
                return (val.hour, val.minute)
    return None


@dataclass
class ClassifiedFile:
    path: Path
    kind: str
    shift_num: Optional[int] = None
    masteryd_direction: Optional[str] = None
    sort_key: Optional[tuple] = None
    masteryd_escale: Optional[str] = None
    masteryd_min_date: Optional[datetime] = None
    masteryd_max_date: Optional[datetime] = None
    masteryd_containers_count: int = 0


def _extract_masteryd_info(header: tuple, rows: list) -> tuple[Optional[str], Optional[datetime], Optional[datetime], int]:
    header_clean = [str(h).strip().upper() if h is not None else "" for h in header]
    
    escale_idx = -1
    date_idx = -1
    for i, h in enumerate(header_clean):
        if "ESCALE" in h or "VOYAGE" in h:
            escale_idx = i
        if "DATE DE SAISIE" in h or "DATE SAISIE" in h:
            date_idx = i
            
    escales = []
    dates = []
    count = 0
    
    for row in rows:
        if not any(c is not None for c in row):
            continue
        count += 1
        if escale_idx != -1 and escale_idx < len(row):
            val = row[escale_idx]
            if val is not None and str(val).strip():
                escales.append(str(val).strip())
        if date_idx != -1 and date_idx < len(row):
            val = row[date_idx]
            if isinstance(val, datetime):
                dates.append(val)
            elif isinstance(val, str):
                try:
                    from dateutil.parser import parse
                    dates.append(parse(val, dayfirst=True))
                except Exception:
                    pass
                    
    major_escale = None
    if escales:
        major_escale = Counter(escales).most_common(1)[0][0]
        
    min_date = min(dates) if dates else None
    max_date = max(dates) if dates else None
    
    return major_escale, min_date, max_date, count


def classify_file(path: Path) -> Optional[ClassifiedFile]:
    if looks_like_template(path):
        logger.info("  %-40s -> TEMPLATE", path.name)
        return ClassifiedFile(path=path, kind="template")
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        logger.warning("Fichier %s illisible, ignoré : %s", path.name, exc)
        return None
    try:
        ws = wb.active
        sheet_title = ws.title
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return None
    if looks_like_shift_report(rows):
        shift_num = detect_shift_number(rows, path, sheet_title)
        sort_key = _earliest_doc_time(rows) or (99, 99)
        if shift_num is not None:
            logger.info("  %-40s -> SHIFT %d", path.name, shift_num)
        else:
            logger.info("  %-40s -> SHIFT (heure DOC=%s)", path.name, sort_key)
        return ClassifiedFile(path=path, kind="shift", shift_num=shift_num, sort_key=sort_key)
    header = rows[0]
    if looks_like_masteryd(header):
        direction = detect_masteryd_direction(header, rows[1:])
        
        # Override/Fallback par préfixe de fichier
        fname_lower = path.name.lower()
        if fname_lower.startswith("imp"):
            direction = "IMPORT"
        elif fname_lower.startswith("exp"):
            direction = "EXPORT"
            
        if direction is None:
            logger.warning("  %-40s -> Direction indéterminée, ignoré.", path.name)
            return None
        
        escale, min_date, max_date, count = _extract_masteryd_info(header, rows[1:])
        logger.info("  %-40s -> %s (Escale: %s)", path.name, direction, escale or "inconnue")
        return ClassifiedFile(
            path=path, kind="masteryd", masteryd_direction=direction,
            masteryd_escale=escale, masteryd_min_date=min_date,
            masteryd_max_date=max_date, masteryd_containers_count=count
        )
    logger.debug("  %-40s -> non reconnu", path.name)
    return None



def discover_input_files(input_dir: Path | None = None,
                          template_dir: Path | None = None,
                          target_escale: str | None = None) -> dict:
    if input_dir is None:
        input_dir = config.INPUT_DIR
    if template_dir is None:
        template_dir = config.TEMPLATE_DIR
    logger.debug("Etape 1 : Decouverte des fichiers par contenu")
    logger.info("Découverte des fichiers d'entrée par contenu dans %s", input_dir)
    search_dirs = [input_dir]
    if template_dir != input_dir:
        search_dirs.append(template_dir)
    candidates = []
    for d in search_dirs:
        if d.exists():
            candidates += [p for p in d.glob("*.xlsx") if not p.name.startswith("~$")]
    classified = [c for c in (classify_file(p) for p in candidates) if c is not None]
    templates = [c for c in classified if c.kind == "template"]
    shifts    = [c for c in classified if c.kind == "shift"]
    masteryds = [c for c in classified if c.kind == "masteryd"]
    
    if not templates:
        raise InputDiscoveryError("Aucun template TPFREP reconnu.")
    if len(templates) > 1:
        logger.warning("Plusieurs templates détectés, on prend le premier.")
    template_path = templates[0].path
    if not shifts:
        raise InputDiscoveryError("Aucun rapport de shift reconnu.")
    
    # 1. Grouper les MASTERYD par escale
    escale_groups: dict[str, list[ClassifiedFile]] = {}
    for m in masteryds:
        escale_name = m.masteryd_escale or "INCONNUE"
        escale_groups.setdefault(escale_name, []).append(m)
        
    # Si plusieurs escales
    if len(escale_groups) > 1 and not target_escale:
        escales_info = []
        for name, files in escale_groups.items():
            min_dates = [f.masteryd_min_date for f in files if f.masteryd_min_date]
            max_dates = [f.masteryd_max_date for f in files if f.masteryd_max_date]
            count = sum(f.masteryd_containers_count for f in files)
            escales_info.append(EscaleInfo(
                name=name,
                min_date=min(min_dates) if min_dates else None,
                max_date=max(max_dates) if max_dates else None,
                container_count=count
            ))
        raise MultipleEscalesError(
            "Plusieurs escales détectées dans data/input. Une sélection est requise.",
            escales=escales_info
        )
        
    # Si un seul groupe ou target_escale précisé
    chosen_escale = target_escale if target_escale else list(escale_groups.keys())[0]
    if chosen_escale not in escale_groups:
        raise InputDiscoveryError(f"L'escale '{chosen_escale}' n'a pas été trouvée dans les fichiers d'entrée.")
        
    valid_masteryds = escale_groups[chosen_escale]
    import_candidates = [m for m in valid_masteryds if m.masteryd_direction == "IMPORT"]
    export_candidates = [m for m in valid_masteryds if m.masteryd_direction == "EXPORT"]
    
    if not import_candidates:
        raise InputDiscoveryError(f"Aucun fichier MASTERYD IMPORT reconnu pour l'escale {chosen_escale}.")
    if not export_candidates:
        raise InputDiscoveryError(f"Aucun fichier MASTERYD EXPORT reconnu pour l'escale {chosen_escale}.")

    # Filtrage des shifts par date et navire
    e_min_dates = [f.masteryd_min_date for f in valid_masteryds if f.masteryd_min_date]
    e_max_dates = [f.masteryd_max_date for f in valid_masteryds if f.masteryd_max_date]
    escale_min_date = min(e_min_dates) if e_min_dates else None
    escale_max_date = max(e_max_dates) if e_max_dates else None
    
    from src.merge import infer_vessel_from_escale, vessel_matches
    # À ce stade les shifts ne sont pas encore nettoyés → infer_vessel_from_escale()
    # avec un dict vide tomberait sur le fallback config (ex. "MASTERY D") et
    # rejetterait tous les shifts d'un autre navire (ex. BELITAKI).
    # On extrait donc le navire directement du nom de l'escale :
    #   "BELITAKI_24062026"    → "BELITAKI"
    #   "MASTERYD_01062026"    → "MASTERYD"
    #   "NORDIC AURORA_..."    → "NORDICAURORA"
    # Le suffixe est toujours _DDMMYYYY (8 chiffres).  En cas d'échec du
    # parsing, on appelle infer_vessel_from_escale comme avant.
    _escale_parts = chosen_escale.rsplit("_", 1)
    if len(_escale_parts) == 2 and _escale_parts[1].isdigit() and len(_escale_parts[1]) == 8:
        target_vessel_norm = _escale_parts[0].replace(" ", "").upper()
    else:
        _tv = infer_vessel_from_escale(chosen_escale, {})
        target_vessel_norm = _tv.replace(" ", "").upper() if _tv else config.TARGET_VESSEL_NORMALIZED
    logger.debug("Filtre navire pour l'escale '%s' : '%s'", chosen_escale, target_vessel_norm)

    margin_days = getattr(config, "SHIFT_DATE_MARGIN_DAYS", 1)
    min_allowed_date = (escale_min_date - timedelta(days=margin_days)) if escale_min_date else None
    max_allowed_date = (escale_max_date + timedelta(days=margin_days)) if escale_max_date else None


    valid_shifts = []
    excluded_shifts = []

    for s in shifts:
        wb = openpyxl.load_workbook(s.path, data_only=True, read_only=True)
        try:
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()
        
        # Extraction de la date
        shift_date = None
        for row in rows[:10]:
            for cell in row:
                if isinstance(cell, datetime):
                    shift_date = cell.replace(hour=0, minute=0, second=0, microsecond=0)
                    break
            if shift_date:
                break
                
        # Filtre de date
        if min_allowed_date and max_allowed_date and shift_date:
            if not (min_allowed_date <= shift_date <= max_allowed_date):
                logger.warning("Fichier %s exclu (Date %s hors de la plage %s - %s)", s.path.name, shift_date.strftime('%Y-%m-%d'), min_allowed_date.strftime('%Y-%m-%d'), max_allowed_date.strftime('%Y-%m-%d'))
                excluded_shifts.append({"file": s.path.name, "reason": "Hors plage de dates de l'escale"})
                continue
                
        # Filtre de navire
        header_idx = None
        for i, row in enumerate(rows[:100]):
            for cell in row:
                if isinstance(cell, str) and "portique" in cell.strip().lower():
                    header_idx = i
                    break
            if header_idx is not None:
                break
                
        vessel_matched = False
        if header_idx is not None:
            subheader_row = rows[header_idx + 1] if header_idx + 1 < len(rows) else []
            header_row = rows[header_idx]
            col_vessel = -1
            for idx in range(max(len(header_row), len(subheader_row))):
                h = str(header_row[idx]).strip().upper() if idx < len(header_row) and header_row[idx] else ""
                sh = str(subheader_row[idx]).strip().upper() if idx < len(subheader_row) and subheader_row[idx] else ""
                combined = f"{h} {sh}"
                if "NAVIRE" in combined or "VESSEL" in combined or "SHIP" in combined or "NOM" in combined:
                    col_vessel = idx
                    break
                    
            if col_vessel != -1:
                for row in rows[header_idx+2:header_idx+30]:
                    if col_vessel < len(row):
                        v_raw = str(row[col_vessel]).strip().upper() if row[col_vessel] else ""
                        if v_raw and vessel_matches(v_raw, target_vessel_norm):
                            vessel_matched = True
                            break
                            
        if not vessel_matched:
            logger.warning("Fichier %s exclu (Aucun mouvement pour le navire cible %s)", s.path.name, target_vessel_norm)
            excluded_shifts.append({"file": s.path.name, "reason": f"Aucun mouvement pour le navire cible ({target_vessel_norm})"})
            continue
            
        valid_shifts.append(s)
        
    shifts = valid_shifts
    if not shifts:
        raise InputDiscoveryError(f"Aucun rapport de shift valide pour l'escale {chosen_escale}.")
    
    # Trier tous les shifts détectés par leur clé chronologique (DOC)
    shifts_sorted = sorted(shifts, key=lambda s: s.sort_key or (99, 99))
    
    # Résoudre ou interpoler les numéros de shifts manquants
    for idx, s in enumerate(shifts_sorted):
        if s.shift_num is None:
            inferred = None
            # Chercher le premier shift numéroté avant dans la liste chronologique
            for prev in reversed(shifts_sorted[:idx]):
                if prev.shift_num is not None:
                    inferred = prev.shift_num + (idx - shifts_sorted.index(prev))
                    break
            if inferred is None:
                # Chercher le premier shift numéroté après dans la liste chronologique
                for nxt in shifts_sorted[idx+1:]:
                    if nxt.shift_num is not None:
                        inferred = nxt.shift_num - (shifts_sorted.index(nxt) - idx)
                        break
            if inferred is None or inferred <= 0:
                inferred = idx + 1
            s.shift_num = inferred
            logger.info("Fichier '%s' assigné au Shift %d par interpolation chronologique.", s.path.name, inferred)
            
    # Créer le dictionnaire final des shifts par numéro, en résolvant les conflits de jours différents
    numbered: dict[int, ClassifiedFile] = {}
    for s in shifts_sorted:
        num = s.shift_num
        if num in numbered:
            # Si deux shifts ont la même heure de DOC (même shift, même jour), c'est un vrai doublon strict
            if s.sort_key == numbered[num].sort_key and s.sort_key != (99, 99):
                raise InputDiscoveryError(
                    f"Conflit de doublons strict : deux fichiers identiques détectés pour le shift {num} "
                    f"({s.path.name} et {numbered[num].path.name})."
                )
            # Sinon, c'est un shift d'un jour suivant (e.g. 'Shift 1' du lendemain), on le renumérote
            logger.info("Décalage chronologique : %s renuméroté de Shift %d vers Shift %d", 
                        s.path.name, num, max(numbered.keys()) + 1)
            num = max(numbered.keys()) + 1
            s.shift_num = num
        
        numbered[num] = s

    shift_nums_sorted = sorted(numbered.keys())
    shift_paths = {f"shift_{n}": numbered[n].path for n in shift_nums_sorted}

    found = {
        **shift_paths,
        "import_masteryd": import_candidates[0].path,
        "export_masteryd": export_candidates[0].path,
        "template":        template_path,
        "_shift_nums":     shift_nums_sorted,
        "_excluded_shifts": excluded_shifts,
        "_chosen_escale": chosen_escale,
    }
    logger.info("Découverte terminée : %d shift(s), IMPORT, EXPORT, template.", len(shift_nums_sorted))
    return found


def load_shift_report(shift_num: int, file_path: Path) -> RawShiftReport:
    logger.info("Lecture Shift %s : %s", shift_num, file_path.name)
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        sheet_title = ws.title
    finally:
        wb.close()
    return RawShiftReport(shift_num=shift_num, file_path=file_path, sheet_title=sheet_title, rows=rows)


def load_masteryd_file(direction: str, file_path: Path) -> RawMasterydFile:
    logger.info("Lecture MASTERYD %s : %s", direction, file_path.name)
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        raise InputDiscoveryError(f"Fichier {file_path.name} vide")
    header = rows[0]
    data_rows = [r for r in rows[1:] if any(c is not None for c in r)]
    return RawMasterydFile(direction=direction, file_path=file_path, header=header, rows=data_rows)


def load_template_workbook(template_path: Path):
    logger.info("Chargement template : %s", template_path.name)
    if not template_path.exists():
        raise InputDiscoveryError(f"Template introuvable : {template_path}")
    return openpyxl.load_workbook(template_path)


@dataclass
class AllInputs:
    shift_reports: dict
    import_masteryd: RawMasterydFile
    export_masteryd: RawMasterydFile
    template_path: Path
    excluded_shifts: list[dict] = field(default_factory=list)
    chosen_escale: Optional[str] = None


def load_all_inputs(input_dir: Path | None = None,
                     template_dir: Path | None = None,
                     target_escale: str | None = None) -> AllInputs:
    if input_dir is None:
        input_dir = config.INPUT_DIR
    if template_dir is None:
        template_dir = config.TEMPLATE_DIR
    paths = discover_input_files(input_dir, template_dir, target_escale)
    shift_nums: list[int] = paths.pop("_shift_nums")
    excluded_shifts = paths.pop("_excluded_shifts", [])
    chosen_escale = paths.pop("_chosen_escale", None)
    
    # On ajoute des infos aux inputs chargés pour les récupérer plus haut
    shift_reports = {n: load_shift_report(n, paths[f"shift_{n}"]) for n in shift_nums}
    import_masteryd = load_masteryd_file("IMPORT", paths["import_masteryd"])
    export_masteryd = load_masteryd_file("EXPORT", paths["export_masteryd"])
    logger.info("Tous les fichiers chargés (%d shift(s)).", len(shift_nums))
    return AllInputs(
        shift_reports=shift_reports,
        import_masteryd=import_masteryd,
        export_masteryd=export_masteryd,
        template_path=paths["template"],
        excluded_shifts=excluded_shifts,    # FIX : valeur maintenant effectivement transmise
        chosen_escale=chosen_escale,        # FIX : idem
    )
